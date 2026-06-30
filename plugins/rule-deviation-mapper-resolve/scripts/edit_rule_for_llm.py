#!/usr/bin/env python3
"""Edit ONLY the `Rule for LLM` cell of given KRI IDs in a Golden-Set xlsx, in place.

Phase-1 / Track-A helper. Never touches any other column or any other rule.
Preserves wrapped/readable view (wrap_text, vertical top); lets row height auto-fit.

Usage (as a library):
    from edit_rule_for_llm import set_rule_for_llm
    set_rule_for_llm("GoldenSet.xlsx", {"SOA-V5-200": "<new rule-for-llm text>", ...})

Notes:
- Finds the `Rule for LLM` column by HEADER NAME (sheets may have extra columns).
- Searches every sheet for each KRI ID (matched in column A).
- Saves in place (overwrites the same file). Caller is responsible for telling the
  user "saved to the file."
"""
import sys, copy
import openpyxl
from openpyxl.styles import Alignment


def _find_col(ws, header="Rule for LLM"):
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == header:
            return c
    return None


def set_rule_for_llm(path, updates: dict, header="Rule for LLM"):
    """updates: {KRI_ID: new_rule_for_llm_text}. Returns list of (sheet,row,kri) updated."""
    wb = openpyxl.load_workbook(path)
    done = []
    remaining = dict(updates)
    for sh in wb.sheetnames:
        ws = wb[sh]
        col = _find_col(ws, header)
        if col is None:
            continue
        for r in range(2, ws.max_row + 1):
            kid = ws.cell(r, 1).value
            if kid in remaining:
                cell = ws.cell(r, col)
                al = copy.copy(cell.alignment)
                cell.value = remaining.pop(kid)
                cell.alignment = Alignment(wrap_text=True, vertical="top",
                                           horizontal=al.horizontal)
                ws.row_dimensions[r].height = None
                done.append((sh, r, kid))
    wb.save(path)
    if remaining:
        print("WARNING: KRI IDs not found:", list(remaining), file=sys.stderr)
    for sh, r, kid in done:
        print(f"updated {sh}!{r}  {kid}")
    return done


if __name__ == "__main__":
    print(__doc__)
