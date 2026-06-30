#!/usr/bin/env python3
"""Append a NEW full-format rule row to a domain sheet of a Golden-Set xlsx (Track B).

Track-B helper. Only use when the protocol grounds the new rule. Copies cell
formatting (alignment/font/border/fill) from the previous data row so the new
row matches the sheet.

Usage (as a library):
    from add_new_rule import add_rule
    add_rule("GoldenSet.xlsx", "OPS", {
        "KRI ID": "OPS72",
        "Category": "Data Handling, Monitoring, and Ethics",
        "KRI Name": "Participant Confidentiality and De-identification",
        "Description": "...",
        "Rule for LLM": "intent: ...\\n...",
        "Protocol Reference & Quote": "Appendix C ... — \\"...\\"",
        "Severity": "major",
        "Deviation Level": "subject",
    })

The values dict is keyed by HEADER NAME; columns are matched by header.
"""
import sys, copy
import openpyxl
from openpyxl.styles import Alignment


def _headers(ws, n=8):
    return [ws.cell(1, c).value for c in range(1, n + 1)]


def add_rule(path, sheet, values: dict):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    headers = _headers(ws, ws.max_column)
    # last data row = highest row with a non-empty column A
    last = max((r for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value), default=1)
    newr = last + 1
    for c, h in enumerate(headers, start=1):
        if h is None:
            continue
        src = ws.cell(last, c)
        dst = ws.cell(newr, c)
        dst.value = values.get(h)
        dst.alignment = copy.copy(src.alignment)
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
    ws.row_dimensions[newr].height = None
    wb.save(path)
    print(f"added {sheet}!{newr}  {values.get('KRI ID')}  |  {values.get('KRI Name')}")
    return newr


if __name__ == "__main__":
    print(__doc__)
