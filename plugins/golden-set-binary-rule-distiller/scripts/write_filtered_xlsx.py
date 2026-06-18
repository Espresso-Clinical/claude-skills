"""
Write the final filtered Golden Set xlsx.

Input JSON is an array of rule objects with keys:
  sheet, KRI ID, Category, KRI Name, Description, Rule for LLM,
  Protocol Reference & Quote, Severity, Deviation Level

The script preserves the source column order, inserts "Deviation Level" right
after "Severity", and applies consistent professional formatting (Arial font,
header fill, wrapped cell text, frozen header row, sensible column widths, and
auto-fit row heights so multi-line YAML cells are fully visible on open).

Usage:
    python write_filtered_xlsx.py <input.json> <output.xlsx>
"""
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


FINAL_COLS = [
    "KRI ID", "Category", "KRI Name", "Description",
    "Rule for LLM", "Protocol Reference & Quote",
    "Severity", "Deviation Level",
]

WIDTHS = {
    "KRI ID": 22, "Category": 22, "KRI Name": 38,
    "Description": 50, "Rule for LLM": 70,
    "Protocol Reference & Quote": 50, "Severity": 11,
    "Deviation Level": 14,
}


def _estimate_row_height(values, widths_by_idx, line_pts=15.0, max_pts=409.0):
    """Approximate auto-fit: the tallest cell drives the row height so wrapped
    multi-line text (notably the YAML Rule for LLM) is fully visible on open."""
    max_lines = 1
    for idx, val in enumerate(values, 1):
        text = "" if val is None else str(val)
        chars_per_line = max(8, int(widths_by_idx.get(idx, 20)))
        lines = 0
        for segment in text.split("\n"):
            seg_len = len(segment.rstrip())
            lines += max(1, -(-seg_len // chars_per_line))  # ceil division
        if lines > max_lines:
            max_lines = lines
    return min(max_pts, max(line_pts, max_lines * line_pts))


def write(in_json: str, out_xlsx: str) -> dict:
    with open(in_json) as f:
        rows = json.load(f)

    by_sheet: dict[str, list[dict]] = {}
    for r in rows:
        by_sheet.setdefault(r.get("sheet", "Sheet1"), []).append(r)

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    body_font = Font(name="Arial")

    counts = {}
    for sheet_name, sheet_rows in by_sheet.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(FINAL_COLS)
        for c in range(1, len(FINAL_COLS) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        for r in sheet_rows:
            ws.append([r.get(c, "") for c in FINAL_COLS])
        for idx, col in enumerate(FINAL_COLS, 1):
            ws.column_dimensions[get_column_letter(idx)].width = WIDTHS[col]
        widths_by_idx = {i: WIDTHS[c] for i, c in enumerate(FINAL_COLS, 1)}
        for ridx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[ridx].height = _estimate_row_height(
                [c.value for c in row], widths_by_idx
            )
        ws.freeze_panes = "A2"
        counts[sheet_name] = len(sheet_rows)

    wb.save(out_xlsx)
    return counts


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(
            "Usage: python write_filtered_xlsx.py <input.json> <output.xlsx>",
            file=sys.stderr,
        )
        sys.exit(2)
    counts = write(sys.argv[1], sys.argv[2])
    print(f"Wrote {sys.argv[2]}")
    for s, n in counts.items():
        print(f"  {s}: {n}")
    print(f"  Total: {sum(counts.values())}")
