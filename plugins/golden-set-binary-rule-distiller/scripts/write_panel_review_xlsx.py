"""
Write a panel-review consolidation xlsx (used for Stage 2 and Stage 4).

Input JSON schema:
{
  "stage": "filter" | "audit",
  "panel_size": 10,
  "reviewers": ["R1", "R2", ..., "R10"],
  "per_reviewer_counts": {"R1": 9, "R2": 11, ...},
  "items": [
    {
      "kri_id": "OPS-052",
      "sheet": "OPS",
      "kri_name": "Same photographer per subject when possible",
      "severity": "minor",
      "votes": 9,
      "votes_by": {"R1": "non-deviation", "R2": "non-deviation", ...},
      "reasons_by": {"R1": "Permissive 'when possible' ...", "R2": "...", ...}
    },
    ...
  ]
}

Produces a workbook with three sheets:
  - Summary       — vote tier counts and recommendations
  - Reviewer Counts — per-reviewer total flag count
  - Consolidated Flags — one row per flagged item, color-coded by tier

Usage:
    python write_panel_review_xlsx.py <input.json> <output.xlsx>
"""
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", start_color="305496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")

TIER_FILLS = {
    10: PatternFill("solid", start_color="006100"),
    9:  PatternFill("solid", start_color="008000"),
    8:  PatternFill("solid", start_color="C6EFCE"),
    7:  PatternFill("solid", start_color="C6EFCE"),
    6:  PatternFill("solid", start_color="FFEB9C"),
    5:  PatternFill("solid", start_color="FFEB9C"),
    4:  PatternFill("solid", start_color="FFE4B5"),
    3:  PatternFill("solid", start_color="FCE4D6"),
    2:  PatternFill("solid", start_color="F2F2F2"),
    1:  PatternFill("solid", start_color="F2F2F2"),
}

REC_BY_TIER = {
    10: "Unanimous — apply",
    9: "Near-unanimous — apply",
    8: "Strong consensus — apply",
    7: "Strong majority — apply",
    6: "Solid majority — apply",
    5: "Half — review",
    4: "Plurality — review (default threshold)",
    3: "Minority — review case-by-case",
    2: "Two votes — likely noise",
    1: "Singleton — likely noise",
}


def write(in_json: str, out_xlsx: str) -> None:
    with open(in_json) as f:
        data = json.load(f)
    panel_size = data.get("panel_size", 10)
    reviewers = data.get("reviewers") or [f"R{i}" for i in range(1, panel_size + 1)]
    counts = data.get("per_reviewer_counts", {})
    items = data.get("items", [])

    # Sort by votes desc, then sheet, then kri_id
    items = sorted(items, key=lambda x: (-x.get("votes", 0), x.get("sheet", ""), x.get("kri_id", "")))

    wb = Workbook()
    wb.remove(wb.active)

    # ----- Summary -----
    ws = wb.create_sheet("Summary")
    ws.append(["Vote tier", "# of items", "Recommendation"])
    for c in range(1, 4):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    tiers = {i: 0 for i in range(1, panel_size + 1)}
    for it in items:
        v = it.get("votes", 0)
        if 1 <= v <= panel_size:
            tiers[v] += 1
    for tier in sorted(tiers.keys(), reverse=True):
        if tiers[tier] > 0:
            ws.append([f"{tier}/{panel_size} reviewers", tiers[tier], REC_BY_TIER.get(tier, "")])
    ws.append(["TOTAL flagged items", sum(tiers.values()), ""])
    for col_letter, w in [("A", 24), ("B", 14), ("C", 38)]:
        ws.column_dimensions[col_letter].width = w
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ----- Reviewer Counts -----
    ws2 = wb.create_sheet("Reviewer Counts")
    ws2.append(["Reviewer", "# of flags raised"])
    for c in range(1, 3):
        cell = ws2.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for r in reviewers:
        ws2.append([r, counts.get(r, 0)])
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 22

    # ----- Consolidated Flags -----
    ws3 = wb.create_sheet("Consolidated Flags")
    cols = (
        ["KRI ID", "Sheet", "KRI Name", "Severity", f"Votes (n/{panel_size})"]
        + list(reviewers)
        + [f"{r} reason" for r in reviewers]
    )
    ws3.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws3.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for it in items:
        row_data = [
            it.get("kri_id", ""), it.get("sheet", ""), it.get("kri_name", ""),
            it.get("severity", ""), it.get("votes", 0),
        ]
        votes_by = it.get("votes_by", {})
        reasons_by = it.get("reasons_by", {})
        for r in reviewers:
            row_data.append(votes_by.get(r, ""))
        for r in reviewers:
            row_data.append(reasons_by.get(r, ""))
        ws3.append(row_data)

    widths = {"KRI ID": 22, "Sheet": 8, "KRI Name": 38, "Severity": 11,
              f"Votes (n/{panel_size})": 11}
    for r in reviewers:
        widths[r] = 13
        widths[f"{r} reason"] = 45
    for idx, c in enumerate(cols, 1):
        ws3.column_dimensions[get_column_letter(idx)].width = widths[c]

    for row_idx, it in enumerate(items, 2):
        fill = TIER_FILLS.get(it.get("votes", 0))
        for col_idx in range(1, len(cols) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill and col_idx <= 5:
                cell.fill = fill
                if it.get("votes", 0) >= 9:
                    cell.font = Font(name="Arial", color="FFFFFF", bold=True)
    ws3.freeze_panes = "F2"
    ws3.row_dimensions[1].height = 36

    wb.save(out_xlsx)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(
            "Usage: python write_panel_review_xlsx.py <input.json> <output.xlsx>",
            file=sys.stderr,
        )
        sys.exit(2)
    write(sys.argv[1], sys.argv[2])
    print(f"Wrote {sys.argv[2]}")
