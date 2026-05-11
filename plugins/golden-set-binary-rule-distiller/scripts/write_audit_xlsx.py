"""
Write the audit / changelog xlsx.

Two modes:
  1. Drop audit  — list of dropped rules with one-line reasons.
  2. Change log  — list of rule modifications with action and one-line reasons.

Input JSON is an array of objects. The script auto-detects the schema based on
the keys present:
  - Drop audit:  Sheet, KRI ID, KRI Name, Category, Severity, Drop Reason
                 (optionally: Panel Votes (n/N))
  - Change log:  Sheet, KRI ID, Action, Note
                 (Action ∈ {DELETED, MODIFIED, ADDED, RENUMBERED, ...})

Usage:
    python write_audit_xlsx.py <input.json> <output.xlsx>
"""
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", start_color="305496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")

ACTION_FILLS = {
    "DELETED": PatternFill("solid", start_color="FCE4D6"),
    "MODIFIED": PatternFill("solid", start_color="FFEB9C"),
    "ADDED": PatternFill("solid", start_color="C6EFCE"),
    "RENUMBERED": PatternFill("solid", start_color="E1D8F1"),
}

DROP_COLS = ["Sheet", "KRI ID", "KRI Name", "Category", "Severity",
             "Panel Votes (n/N)", "Drop Reason"]
DROP_WIDTHS = {"Sheet": 10, "KRI ID": 24, "KRI Name": 42, "Category": 24,
               "Severity": 11, "Panel Votes (n/N)": 14, "Drop Reason": 70}

CHANGE_COLS = ["Sheet", "KRI ID", "Action", "Note"]
CHANGE_WIDTHS = {"Sheet": 10, "KRI ID": 26, "Action": 14, "Note": 90}


def detect_schema(rows):
    if not rows:
        return "drop"  # default
    keys = set(rows[0].keys())
    if "Action" in keys:
        return "change"
    return "drop"


def write(in_json: str, out_xlsx: str) -> None:
    with open(in_json) as f:
        rows = json.load(f)
    schema = detect_schema(rows)
    cols = CHANGE_COLS if schema == "change" else DROP_COLS
    widths = CHANGE_WIDTHS if schema == "change" else DROP_WIDTHS

    wb = Workbook()
    ws = wb.active
    ws.title = "Changelog" if schema == "change" else "Dropped Rules"
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    for idx, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths[col]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        action = ""
        if schema == "change":
            action = (rows[row_idx - 2].get("Action") or "").upper()
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if schema == "change" and cell.column == 3 and action in ACTION_FILLS:
                cell.fill = ACTION_FILLS[action]
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    wb.save(out_xlsx)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(
            "Usage: python write_audit_xlsx.py <input.json> <output.xlsx>",
            file=sys.stderr,
        )
        sys.exit(2)
    write(sys.argv[1], sys.argv[2])
    print(f"Wrote {sys.argv[2]}")
