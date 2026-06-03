"""Stage 4 — quality audit panel on Gemini 3.5 Flash (high thinking).
3 reviewers per chunk audit the authored YAML rules against their source rows.
Consensus flag at >=2/3. Categories: omitted-detail, jargon, filler, bad-denominator,
restated-criterion, names-data, unclear, non-deviation, wrong-level, fabricated-specifics.
"""
import sys, os, json, re
from concurrent.futures import ThreadPoolExecutor
from collections import defaultdict, Counter
SOA = os.path.expanduser("~/.claude/skills-repo/plugins/soa-kri-extractor/scripts")
sys.path.insert(0, SOA)
from gemini_extract import call_gemini
import openpyxl

RUN = "/Users/ofir/Downloads/extractor/ENX-CL-05-001/run_002/binary_distill"
XLSX = "/Users/ofir/Downloads/ENX-CL-05-001_rules for LLM.xlsx"
kd = json.load(open(f"{RUN}/stage2_keepdrop.json"))
survivors = [r["kri_id"] for r in kd if r["keep"]]
authored = json.load(open(f"{RUN}/authored_rules.json"))

wb = openpyxl.load_workbook(XLSX, data_only=True)
detail = {}
for sh in ["ELIG","SAF","END","OPS","SOA"]:
    ws = wb[sh]; hdr={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)}
    for r in range(2, ws.max_row+1):
        kid = ws.cell(r,1).value
        if not kid: continue
        detail[kid] = {"sheet":sh,"name":ws.cell(r,hdr["KRI Name"]).value,
            "desc":ws.cell(r,hdr["Description"]).value,"ref":ws.cell(r,hdr["Protocol Reference & Quote"]).value}

HEADER = '''You are an INDEPENDENT QC reviewer auditing authored "Protocol rules" (YAML) for a clinical-trial KRI engine. Each rule below shows its SOURCE (name/description/protocol reference+quote) and the AUTHORED YAML. Flag problems.

FLAG CATEGORIES:
- omitted-detail: a checkable detail in the source/quote/footnote (analyte, time/acceptability window, required subset, AND/OR logic, conditional trigger, pass/deviation condition) is MISSING from the YAML. (Most important.)
- jargon: meta-commentary or tags instead of plain clinical language (e.g. "presence and dating only - does NOT require X", "[footnote N definition]").
- filler: boilerplate adding length without a check ("active subject expected to attend...").
- bad-denominator: applies_to is a data filter or wrong population ("randomized" for an eligibility criterion, "active").
- restated-criterion: evidence_expected just restates the criterion instead of naming a clinical artifact.
- names-data: rule names a table/column/code/field/join (must be data-agnostic).
- unclear: a slot is vague or missing data the engine needs.
- non-deviation: the deviation slot is not a real anomaly.
- wrong-level: (we also pass a deviation_level guess) — ignore unless clearly wrong.
- fabricated-specifics: YAML introduces a number/tolerance/item NOT in the source.

NOTE: SOA rules are templated presence/timing checks - do not nitpick them uniformly; flag only genuine problems.

OUTPUT: ONLY strict JSON (no prose, no fences): a list of flagged rules:
[{"kri_id":"...","flags":["omitted-detail", ...],"note":"<=20 words"}]
Return [] if a chunk has no problems.

=== RULES ===
'''

def rule_text(kid):
    d = detail[kid]
    return (f'--- {kid} [{d["sheet"]}] ---\nSOURCE name: {d["name"]}\nSOURCE desc: {d["desc"]}\n'
            f'SOURCE ref/quote: {d["ref"]}\nAUTHORED:\n{authored.get(kid,"<missing>")}\n')

CHUNK=40
chunks=[survivors[i:i+CHUNK] for i in range(0,len(survivors),CHUNK)]

def parse_json(txt):
    t=txt.strip(); t=re.sub(r"^```(json)?","",t).strip(); t=re.sub(r"```$","",t).strip()
    m=re.search(r"\[.*\]",t,re.DOTALL)
    if m: t=m.group(0)
    try: return json.loads(t)
    except Exception: return None

temps=[0.1,0.2,0.3]
def review(args):
    ci,ri=args
    prompt=HEADER+"\n".join(rule_text(k) for k in chunks[ci])
    try:
        out=call_gemini(prompt, system_prompt=f"QC reviewer #{ri+1} of 3. Be rigorous and independent.",
                        temperature=temps[ri], task="judge")
        return ci,ri,parse_json(out)
    except Exception as e:
        return ci,ri,{"error":str(e)}

jobs=[(ci,ri) for ci in range(len(chunks)) for ri in range(3)]
votes=defaultdict(Counter)   # kid -> Counter(category -> n reviewers)
notes=defaultdict(list)
errors=0
with ThreadPoolExecutor(max_workers=6) as ex:
    for ci,ri,res in ex.map(review, jobs):
        if isinstance(res,list):
            for item in res:
                kid=item.get("kri_id")
                if kid in authored:
                    for fl in item.get("flags",[]):
                        votes[kid][fl]+=1
                    if item.get("note"): notes[kid].append(item["note"])
        else:
            errors+=1

# consensus: a (kid,flag) with >=2 of 3 reviewers in its chunk
consensus=[]
for kid,cnt in votes.items():
    cflags=[f for f,n in cnt.items() if n>=2]
    if cflags:
        consensus.append({"kri_id":kid,"sheet":detail[kid]["sheet"],"flags":cflags,
                          "votes":dict(cnt),"notes":notes.get(kid,[])[:2]})
consensus.sort(key=lambda x:x["kri_id"])
json.dump({"errors":errors,"consensus_flags":consensus,
           "all_votes":{k:dict(v) for k,v in votes.items()}},
          open(f"{RUN}/stage4_panel.json","w"), indent=1, ensure_ascii=False)

bycat=Counter()
for c in consensus:
    for f in c["flags"]: bycat[f]+=1
print(f"Stage 4 audit: {len(survivors)} rules, {len(jobs)} reviewer-passes, {errors} errors")
print(f"Rules with >=2/3 consensus flag: {len(consensus)}")
print("By category:", dict(bycat))
for c in consensus[:60]:
    print(f"  {c['kri_id']:18} {c['flags']}  {c['notes'][0] if c['notes'] else ''}")
print("saved:", f"{RUN}/stage4_panel.json")
