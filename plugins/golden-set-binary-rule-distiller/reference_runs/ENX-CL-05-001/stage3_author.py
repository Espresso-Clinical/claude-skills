"""Stage 3 — author YAML Protocol rules for the 298 survivors, on Gemini 3.5 Flash (high thinking).
Batched + threaded. Returns @@@KRI/@@@END blocks, parsed and YAML-validated.
"""
import sys, os, json, re, yaml
from concurrent.futures import ThreadPoolExecutor
SOA = os.path.expanduser("~/.claude/skills-repo/plugins/soa-kri-extractor/scripts")
sys.path.insert(0, SOA)
from gemini_extract import call_gemini
import openpyxl

RUN = "/Users/ofir/Downloads/extractor/ENX-CL-05-001/run_002/binary_distill"
XLSX = "/Users/ofir/Downloads/ENX-CL-05-001_rules for LLM.xlsx"

kd = json.load(open(f"{RUN}/stage2_keepdrop.json"))
survivors = [r for r in kd if r["keep"]]

wb = openpyxl.load_workbook(XLSX, data_only=True)
detail = {}
for sh in ["ELIG","SAF","END","OPS","SOA"]:
    ws = wb[sh]; hdr={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)}
    for r in range(2, ws.max_row+1):
        kid = ws.cell(r,1).value
        if not kid: continue
        detail[kid] = {"sheet":sh,
            "name":ws.cell(r,hdr["KRI Name"]).value,
            "desc":ws.cell(r,hdr["Description"]).value,
            "ref":ws.cell(r,hdr["Protocol Reference & Quote"]).value}

SPEC = '''You author "Protocol rules" for a clinical-trial KRI monitoring engine. For EACH rule below, write a YAML "Protocol rule": a DATA-AGNOSTIC clinical statement of the check.

SLOTS (YAML, in this order):
intent: one-line plain-English purpose
applies_to: the clinical denominator (WHO/WHAT must satisfy this). For ELIGIBILITY criteria use "enrolled subjects" (assessed at screening; the deviation is enrolling an ineligible subject). NEVER "randomized", NEVER "active", never a data filter. Use clinical denominators like "every SAE", "every activated site", "every IP administration".
evidence_expected: the clinical artifact that must exist (a history / assessment / finding / result / record / log). NEVER just restate the criterion.
acceptance: a mapping with an OPEN set of sub-slots - use only those the rule needs: timing (window/acceptability window), required (mandatory items), preferred (non-mandatory but preferred), conditional (trigger/exception), pass (plain pass condition), override (documented waiver).
deviation: the violation in clinical terms (derived from applies_to + acceptance)
provenance: terse pointer ONLY - section + footnote(s) + page (do NOT repeat the full verbatim quote)

HARD RULES:
- DATA-AGNOSTIC: never name tables, columns, codes, EDC/CRF fields, or joins.
- COMPREHENSIVE: fold in EVERY checkable detail from the rule's Name, Description and Reference/Quote (including any footnote text): analyte/parameter lists, time & acceptability windows, required subsets with AND/OR logic, conditional triggers, pass/deviation conditions. Never leave a checkable detail out.
- CLARITY: plain clinical language only. NO meta-commentary like "presence and dating only - does NOT require X" and NO tags like "[footnote 12 definition]".
- PRESENCE-ONLY LABS: for a SCREENING lab, put the component list inside evidence_expected as a description, give the window in acceptance.timing, and make deviation = "no result performed/dated within the window" - do NOT add a `required` item list. For a PRE-TREATMENT lab that a footnote makes mandatory, DO use `required` (the mandatory subset, with AND/OR) + `preferred` (the full panel).
- NEVER invent thresholds, tolerances, or items not in the source; preserve fuzzy wording ("approximately").

EXAMPLES:
@@@KRI: EXAMPLE-required-subset-lab
intent: "Pre-treatment biochemistry was available and reviewed before the first dose."
applies_to: "subjects who reached the first treatment"
evidence_expected: "a biochemistry result usable for the pre-treatment review."
acceptance:
  timing: "drawn up to 4 days before the visit and reviewed before first treatment [footnote 14]"
  required: "sodium, potassium, glucose, bilirubin, creatinine, (AST or ALT) [footnote 14 mandatory]"
  preferred: "full panel (albumin, LDH, GGT, ALP, CRP/hsCRP, urea, total protein)"
  conditional: "direct bilirubin only if total bilirubin is abnormal"
deviation: "a subject with no qualifying biochemistry in the window carrying the required items."
provenance: "SoA Run-In, Footnote 14, p.26."
@@@END
@@@KRI: EXAMPLE-presence-only-lab
intent: "A complete blood count was performed for the Screening visit."
applies_to: "subjects who attended the Screening visit"
evidence_expected: "a complete blood count result for Screening (RBC, hemoglobin, hematocrit, MCH, MCHC, MCV, WBC with differential, platelets)."
acceptance:
  timing: "performed and dated for Screening; a result from up to 3 months before eligibility confirmation is acceptable [footnote 13]"
deviation: "a Screening attendee with no complete blood count performed and dated within the acceptable window."
provenance: "SoA Run-In, Footnotes 12 & 13, p.26."
@@@END
@@@KRI: EXAMPLE-eligibility-pass-override
intent: "Subject had no other IP / interventional-trial participation within 60 days of first treatment."
applies_to: "enrolled subjects"
evidence_expected: "the subject's prior investigational-product / interventional-trial exposure history for the 60 days before first treatment, and any documented Sponsor approval."
acceptance:
  pass: "no IP receipt or interventional-trial participation in the 60 days before first treatment"
  override: "documented Sponsor approval waives the criterion"
deviation: "an enrolled subject who failed this without documented Sponsor approval."
provenance: "Section 8.2 p.62, Exclusion #22."
@@@END

OUTPUT FORMAT: for EACH rule below output a block exactly:
@@@KRI: <kri_id>
<the YAML rule>
@@@END
Output nothing else - no prose, no fences.

=== RULES TO AUTHOR ===
'''

def rule_block(kid):
    d = detail[kid]
    return f'[{d["sheet"]}] {kid}\n  Name: {d["name"]}\n  Description: {d["desc"]}\n  Reference/Quote: {d["ref"]}\n'

# batch (keep sheets together; ~10 per batch)
batches = []
for sh in ["ELIG","SAF","END","OPS","SOA"]:
    ids = [r["kri_id"] for r in survivors if r["sheet"]==sh]
    for i in range(0, len(ids), 10):
        batches.append(ids[i:i+10])

def author_batch(bi):
    ids = batches[bi]
    prompt = SPEC + "\n".join(rule_block(k) for k in ids)
    try:
        out = call_gemini(prompt, system_prompt="You are a meticulous clinical protocol-rule author. Follow the format exactly.",
                          temperature=0.15, task="judge")
        return bi, out
    except Exception as e:
        return bi, f"__ERROR__ {e}"

raw = {}
with ThreadPoolExecutor(max_workers=6) as ex:
    for bi, out in ex.map(author_batch, range(len(batches))):
        raw[bi] = out

# parse @@@KRI blocks
authored = {}
block_re = re.compile(r"@@@KRI:\s*([^\n]+?)\s*\n(.*?)@@@END", re.DOTALL)
for bi, out in raw.items():
    if out.startswith("__ERROR__"):
        print(f"batch {bi} ERROR: {out[:100]}"); continue
    for m in block_re.finditer(out):
        kid = m.group(1).strip()
        body = m.group(2).strip()
        body = re.sub(r"^```(yaml)?","",body).strip(); body=re.sub(r"```$","",body).strip()
        authored[kid] = body

# validate
REQ = ["intent","applies_to","evidence_expected","acceptance","deviation","provenance"]
ok=0; bad=[]
for kid in [r["kri_id"] for r in survivors]:
    y = authored.get(kid)
    if not y: bad.append((kid,"MISSING")); continue
    try:
        p = yaml.safe_load(y)
        miss=[s for s in REQ if s not in p]
        if miss or not isinstance(p.get("acceptance"),dict) or not p["acceptance"]:
            bad.append((kid,f"slots:{miss or 'acceptance-empty'}"))
        else: ok+=1
    except Exception as e:
        bad.append((kid,f"yaml-error:{str(e)[:40]}"))

json.dump(authored, open(f"{RUN}/authored_rules.json","w"), indent=1, ensure_ascii=False)
print(f"Authored OK: {ok}/{len(survivors)}")
print(f"Problems: {len(bad)}")
for kid,why in bad[:40]: print(f"  {kid}: {why}")
print("saved:", f"{RUN}/authored_rules.json")
