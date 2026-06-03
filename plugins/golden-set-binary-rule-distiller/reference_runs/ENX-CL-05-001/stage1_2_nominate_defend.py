"""Re-run the binary FILTER under the new keep-biased Nominate -> Defend design.
Stage 1 (nominate) + Stage 2 (defend) + inverse-coverage, all Gemini 3.5 Flash high thinking.
Compares the result against run_002's 37 drops (esp. the 6 governance rules)."""
import sys, os, json, re, threading
SOA=os.path.expanduser("~/.claude/skills-repo/plugins/soa-kri-extractor/scripts"); sys.path.insert(0,SOA)
from gemini_extract import call_gemini
import openpyxl
RUN="/Users/ofir/Downloads/extractor/ENX-CL-05-001/run_003/binary_distill"
XLSX="/Users/ofir/Downloads/ENX-CL-05-001_rules for LLM.xlsx"
OLD=json.load(open("/Users/ofir/Downloads/extractor/ENX-CL-05-001/run_002/binary_distill/stage2_keepdrop.json"))
old_drops={r['kri_id']:r for r in OLD if not r['keep']}

wb=openpyxl.load_workbook(XLSX,data_only=True); rules=[]
for sh in ['ELIG','SAF','END','OPS','SOA']:
    ws=wb[sh]; h={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)}
    for r in range(2,ws.max_row+1):
        k=ws.cell(r,1).value
        if not k: continue
        ref=str(ws.cell(r,h['Protocol Reference & Quote']).value or '')[:280]
        rules.append({'kri_id':k,'sheet':sh,'name':ws.cell(r,h['KRI Name']).value or '',
                      'desc':str(ws.cell(r,h['Description']).value or '')[:200],'ref':ref})
def line(r): return f"{r['kri_id']} | {r['sheet']} | {r['name']} :: {r['desc']} :: REF {r['ref']}"
ALL="\n".join(line(r) for r in rules)

def parse(txt):
    t=txt.strip(); t=re.sub(r"^```(json)?","",t).strip(); t=re.sub(r"```$","",t).strip()
    m=re.search(r"\[.*\]",t,re.DOTALL); t=m.group(0) if m else t
    try: return json.loads(t)
    except: return None
def panel(prompt,sysp,n=5):
    temps=[0.1,0.2,0.3,0.15,0.25]; res=[None]*n
    def run(i):
        try: res[i]=parse(call_gemini(prompt,system_prompt=sysp.format(i=i+1,n=n),temperature=temps[i%len(temps)],task="judge"))
        except Exception as e: res[i]={"error":str(e)}
    ts=[threading.Thread(target=run,args=(i,)) for i in range(n)]; [t.start() for t in ts]; [t.join() for t in ts]
    return res

# ---- STAGE 1: nominate ----
S1=f"""You are reviewer #{{i}} of {{n}} screening a clinical-trial Golden Set for rules that CANNOT be turned into a binary, deviation-producing rule.
KEEPABLE = a binary check can be authored (measurable anchor: number, date, presence/absence, category; fuzzy thresholds are fine).
NOMINATE FOR DROP only if primarily: pure endpoint/population definition, statistical methodology, reporting metric, exploratory analysis, permissive "may" option, broad meta-compliance, pure term/threshold definition, pure aspiration, or purely subjective investigator judgment with no objective anchor.
KEEP IS THE DEFAULT - when unsure, do NOT nominate. SOA visit x procedure rules are templated; nominate only genuinely optional/conditional/unscheduled ones.
OUTPUT ONLY strict JSON: [{{"kri_id":"...","drop_reason":"<=12 words"}}]  ([] if none).

RULES:
{ALL}"""
s1=panel(S1,"You screen clinical KRIs. Reviewer #{i}/{n}. Keep-biased, independent.")
from collections import Counter
nom=Counter(); nom_reason={}
for res in s1:
    if isinstance(res,list):
        for it in res:
            k=it.get('kri_id')
            if any(r['kri_id']==k for r in rules): nom[k]+=1; nom_reason.setdefault(k,it.get('drop_reason',''))
candidates=sorted(nom)
json.dump({'nominations':dict(nom),'reasons':nom_reason,'candidates':candidates},open(f"{RUN}/stage1_nominations.json","w"),indent=1,ensure_ascii=False)
print(f"STAGE 1: {len(candidates)} drop-candidates nominated (of {len(rules)} rules).")

# ---- STAGE 2: defend ----
det={r['kri_id']:r for r in rules}
cand_block="\n".join(line(det[k]) for k in candidates)
S2=f"""You are reviewer #{{i}} of {{n}}. The rules below were NOMINATED to drop as "not binary". Your job is the OPPOSITE: genuinely try to DEFEND keeping each.
For each, verdict:
- "defend": there IS a protocol-grounded way to author a binary, deviation-producing rule - a measurable anchor, a footnote making it testable, OR a legitimate trial-level/site-level governance KRI (interim/enrollment trigger, cohort gating, DSMB approval, randomization stratification, IRB approval, IP accountability). Give the basis.
- "confirm": after genuinely trying, there is no objective checkable anchor - truly definitional/permissive/subjective.
BIAS TOWARD "defend". Only "confirm" if you find no protocol basis to keep it.
OUTPUT ONLY strict JSON: [{{"kri_id":"...","verdict":"defend"|"confirm","basis":"<=18 words"}}]

CANDIDATES:
{cand_block}"""
s2=panel(S2,"You defend clinical KRIs against removal. Reviewer #{i}/{n}. Independent.")
defend=Counter(); confirm=Counter(); basis={}
for res in s2:
    if isinstance(res,list):
        for it in res:
            k=it.get('kri_id'); v=it.get('verdict')
            if k not in candidates: continue
            if v=='defend': defend[k]+=1; basis.setdefault(k,it.get('basis',''))
            elif v=='confirm': confirm[k]+=1
verdict={}
for k in candidates:
    c,d=confirm[k],defend[k]
    if c>=4 and d==0: verdict[k]='DROP'
    elif d>=3: verdict[k]='RESTORE'
    else: verdict[k]='ESCALATE'
drops=[k for k in candidates if verdict[k]=='DROP']
json.dump({'confirm':dict(confirm),'defend':dict(defend),'basis':basis,'verdict':verdict},open(f"{RUN}/stage2_defense.json","w"),indent=1,ensure_ascii=False)

# ---- inverse coverage ----
drop_block="\n".join(f"{k} | {det[k]['sheet']} | {det[k]['name']} (confirm {confirm[k]}/5)" for k in drops)
IC=f"""Below is the FULL finalized DROP list from a binary filter. Review as a SET. Identify any rule or whole CLASS wrongly removed - especially legitimate trial-level/site-level governance KRIs or anything with a real protocol-grounded data check.
OUTPUT ONLY strict JSON: [{{"kri_id":"...","why_keep":"<=18 words"}}]  ([] if clean).

DROPPED:
{drop_block}"""
ic=parse(call_gemini(IC,system_prompt="You audit a drop list for over-dropping.",temperature=0.1,task="judge")) or []
ic_ids={x.get('kri_id') for x in ic if isinstance(x,dict)}
json.dump(ic,open(f"{RUN}/inverse_coverage.json","w"),indent=1,ensure_ascii=False)

final_drops=[k for k in drops if k not in ic_ids]
print(f"STAGE 2: DROP {sum(v=='DROP' for v in verdict.values())} | RESTORE {sum(v=='RESTORE' for v in verdict.values())} | ESCALATE {sum(v=='ESCALATE' for v in verdict.values())}")
print(f"INVERSE-COVERAGE restored {len(ic_ids)}: {sorted(ic_ids)}")
print(f"FINAL drops: {len(final_drops)}  (run_002 dropped 37)")

# ---- comparison vs old 37 ----
print("\n=== How the OLD 37 drops fare under Nominate->Defend ===")
gov={'GOV-INTER-001','GOV-INTER-002','GOV-STAT-003','GOV-ASET-001','END-OBL-032','OPS-COMP-001'}
for k in sorted(old_drops):
    if k in verdict: v=verdict[k]; vc=f"(confirm {confirm[k]}/5, defend {defend[k]}/5)"
    else: v='NOT-NOMINATED (kept)'; vc=''
    if k in ic_ids: v+=' +IC-restore'
    star=' <== GOVERNANCE' if k in gov else ''
    print(f"  {k:24} {v:22} {vc}{star}")
newly=[k for k in final_drops if k not in old_drops]
print("\nNEW drops not in run_002:", newly if newly else "none")
json.dump({'final_drops':final_drops,'restored_or_escalated':[k for k in candidates if verdict[k]!='DROP'],
           'inverse_restored':sorted(ic_ids)},open(f"{RUN}/refilter_result.json","w"),indent=1,ensure_ascii=False)
print("\nsaved ->",RUN)
