"""Finalize: apply Stage-4 content fixes + strip footnote NUMBERS (keep content) from all
298 authored rules, then assemble the final distilled xlsx + drop audit + changelog."""
import json, re, collections, openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import yaml

RUN = "/Users/ofir/Downloads/extractor/ENX-CL-05-001/run_002/binary_distill"
XLSX = "/Users/ofir/Downloads/ENX-CL-05-001_rules for LLM.xlsx"
authored = json.load(open(f"{RUN}/authored_rules.json"))
kd = json.load(open(f"{RUN}/stage2_keepdrop.json"))
survivors = [r for r in kd if r["keep"]]
level = {r["kri_id"]: r["level_hint"] for r in kd}

def OD(*kv):
    d=collections.OrderedDict()
    for k,v in kv: d[k]=v
    return d

# ---- hand-authored corrections (Stage-4 content fixes), already footnote-number-free ----
OV = {}
OV['SOA-061']=OD(('intent','The Screening visit occurred within its protocol window before the first treatment.'),
 ('applies_to','subjects who attended the Screening visit'),
 ('evidence_expected','the Screening visit date relative to the first treatment (Day 0).'),
 ('acceptance',OD(('timing','Screening performed within Day -29 to 0 (up to 29 days before, and before, the first treatment)'))),
 ('deviation','a subject whose Screening visit fell outside the Day -29 to 0 window.'),
 ('provenance','Schedule of Activities, p.24-34.'))
OV['SOA-103']=OD(('intent','Informed consent was in place before any randomization-phase screening procedure.'),
 ('applies_to','subjects who attended the Randomization-phase Screening visit'),
 ('evidence_expected','a signed and dated informed consent form for the randomization phase.'),
 ('acceptance',OD(('timing','signed before any randomization-phase study procedure'),
   ('conditional','the patient is re-consented if more than 60 days passed between consent and the first treatment'))),
 ('deviation','a randomization-phase screening subject with no signed and dated informed consent before procedures, or not re-consented when more than 60 days elapsed before first treatment.'),
 ('provenance','Schedule of Activities (Randomized Phase), p.31-34.'))
OV['SOA-104']=OD(('intent','Eligibility (inclusion/exclusion) was assessed at the randomization-phase Screening visit.'),
 ('applies_to','subjects who attended the Randomization-phase Screening visit'),
 ('evidence_expected','the inclusion/exclusion eligibility assessment recorded at the randomization-phase Screening visit.'),
 ('acceptance',OD(('timing','performed at the Randomization-phase Screening visit'))),
 ('deviation','a randomization-phase screening subject with no inclusion/exclusion eligibility assessment recorded at that visit.'),
 ('provenance','Schedule of Activities (Randomized Phase), p.31-34.'))
OV['SAF-PREG-002']=OD(('intent','Each reported pregnancy was followed to outcome and the outcome notified to the Sponsor.'),
 ('applies_to',"pregnancies in a study patient or a study patient's partner reported after enrollment"),
 ('evidence_expected','the pregnancy follow-up record through completion or termination, and the outcome notification to the Sponsor.'),
 ('acceptance',OD(('required','the pregnancy is followed until completion or termination (e.g. induced abortion) and the outcome is notified to the Sponsor'))),
 ('deviation','a reported pregnancy not followed to completion/termination, or whose outcome was not notified to the Sponsor.'),
 ('provenance','Section 15.9, p.85.'))
OV['ELIG-INC-004d']=OD(('intent','Across the screening pain diary, the spread between the lowest and highest daily index-knee pain score was at most 3 points.'),
 ('applies_to','enrolled subjects'),
 ('evidence_expected','the screening index-knee daily pain diary (NRS 0-10 over 10 days, at least 5 measurements after at least 48 hours of analgesic washout).'),
 ('acceptance',OD(('pass','no more than 3 points between the lowest and highest daily pain score'))),
 ('deviation','an enrolled subject whose screening daily pain scores spanned more than 3 points between lowest and highest.'),
 ('provenance','Section 8.1, p.59, Inclusion #4d.'))
OV['SOA-199']=OD(('intent','Screening was completed within its window and only after informed consent.'),
 ('applies_to','subjects who attended Screening'),
 ('evidence_expected','the informed-consent date and the screening dates relative to the first IP administration (Day 0).'),
 ('acceptance',OD(('timing','screening carried out between Day -29 and Day 0, following signing of informed consent and before the first IP administration'))),
 ('deviation','a subject screened outside Day -29 to 0, or screened before informed consent was signed.'),
 ('provenance','Section 7.1, p.55.'))
def biochem(visit, label):
    return OD(('intent',f'Pre-treatment biochemistry was available and reviewed before the {label} dose.'),
     ('applies_to',f'subjects who reached the {visit} visit'),
     ('evidence_expected',f'a biochemistry result usable for the {visit} pre-treatment review.'),
     ('acceptance',OD(('timing','drawn up to one week before the visit and reviewed before the treatment injection'),
       ('required','sodium, potassium, glucose, bilirubin, creatinine, and (AST or ALT)'))),
     ('deviation',f'a {visit} subject with no qualifying biochemistry in the window carrying the required items reviewed before the treatment injection.'),
     ('provenance','Schedule of Activities (Run-In Phase), p.26.'))
OV['SOA-009']=biochem('V3 (Treatment #2)','V3 (Treatment #2, Day 12-16)')
OV['SOA-010']=biochem('V5 (Treatment #3)','V5 (Treatment #3, Day 26-30)')
def simplelab(visit, what, evid):
    return OD(('intent',f'A {what} was available for the {visit} visit.'),
     ('applies_to',f'subjects who reached the {visit} visit'),
     ('evidence_expected',evid),
     ('acceptance',OD(('timing','performed and dated for the visit; may be drawn up to one week before the visit'))),
     ('deviation',f'a {visit} subject with no {what} result within the acceptable window.'),
     ('provenance','Schedule of Activities (Run-In Phase), p.26.'))
OV['SOA-012']=simplelab('V3 (Treatment #2)','coagulation test','a coagulation test result for the V3 visit.')
OV['SOA-013']=simplelab('V5 (Treatment #3)','coagulation test','a coagulation test result for the V5 visit.')
OV['SOA-016']=simplelab('V3 (Treatment #2)','complete blood count','a complete blood count result for the V3 visit (red blood cell count, hemoglobin, hematocrit, MCH, MCHC, MCV, white blood cell count with differential, and platelet count).')
OV['SOA-017']=simplelab('V5 (Treatment #3)','complete blood count','a complete blood count result for the V5 visit.')

def q(s):
    s=str(s).replace('\\','\\\\').replace('"','\\"'); return '"'+s+'"'
def to_yaml(d):
    out=[]
    for k,v in d.items():
        if isinstance(v,dict):
            out.append(f"{k}:")
            for sk,sv in v.items(): out.append(f"  {sk}: {q(sv)}")
        else: out.append(f"{k}: {q(v)}")
    return "\n".join(out)

# ---- footnote-number stripper (operate on parsed values, re-emit) ----
def clean(s):
    s=str(s)
    s=re.sub(r"\s*\[[^\]]*[Ff]ootnote[^\]]*\]","",s)                       # [footnote 13], [footnote 12 & 13]
    s=re.sub(r"\b[Ff]ootnotes?\s*\d+(?:\s*(?:&|and|,)\s*\d+)*","",s)        # Footnote 14 / Footnotes 12 & 13
    s=re.sub(r"\b[Ff]ootnotes?\s*\d+","",s)                                 # any leftover single
    s=re.sub(r",?\s*&\s*,",",",s); s=re.sub(r",\s*,",",",s)
    s=re.sub(r"\(\s*\)","",s); s=re.sub(r"\s*&\s*(?=[,\.])","",s)
    s=re.sub(r"\s+([,.;])",r"\1",s); s=re.sub(r"[ \t]{2,}"," ",s)
    s=re.sub(r",\s*\.",".",s)
    s=s.replace(" -  "," - ").strip().strip(",").strip()
    return s
def clean_rule(yamlstr):
    p=yaml.safe_load(yamlstr)
    out=collections.OrderedDict()
    for k,v in p.items():
        if isinstance(v,dict):
            out[k]=collections.OrderedDict((sk,clean(sv)) for sk,sv in v.items())
        else: out[k]=clean(v)
    return out

final={}
changelog=[]
for r in survivors:
    kid=r["kri_id"]
    if kid in OV:
        d=OV[kid]; changelog.append((kid,"content-fix + footnote-strip","Stage-4 correction"))
    else:
        d=clean_rule(authored[kid])
    final[kid]=to_yaml(d)

# validate
REQ=["intent","applies_to","evidence_expected","acceptance","deviation","provenance"]
bad=[]; leftover=[]
for kid in [r["kri_id"] for r in survivors]:
    y=final[kid]
    try:
        p=yaml.safe_load(y)
        if [s for s in REQ if s not in p] or not isinstance(p.get("acceptance"),dict) or not p["acceptance"]:
            bad.append(kid)
    except Exception as e: bad.append(f"{kid}:{e}")
    if re.search(r"[Ff]ootnote\s*\d",y): leftover.append(kid)
print("validate: rules", len(final), "| slot/yaml problems", len(bad), bad[:10])
print("footnote-number leftovers:", len(leftover), leftover[:10])

# ---- assemble final xlsx ----
src=openpyxl.load_workbook(XLSX,data_only=True)
SRC=['KRI ID','Category','KRI Name','Description','Protocol Reference & Quote','Severity','Deviation Level','Arm','Cohort']
OUT=['KRI ID','Category','KRI Name','Description','Rule for LLM','Protocol Reference & Quote','Severity','Deviation Level','Arm','Cohort']
keep_ids={r["kri_id"] for r in survivors}
wbo=Workbook(); wbo.remove(wbo.active)
for sh in ['ELIG','SAF','END','OPS','SOA']:
    ws=src[sh]; h={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)}
    o=wbo.create_sheet(sh)
    for c,col in enumerate(OUT,1): o.cell(1,c,col).font=Font(bold=True)
    ro=2
    for r in range(2,ws.max_row+1):
        kid=ws.cell(r,1).value
        if kid not in keep_ids: continue
        vals={col:ws.cell(r,h[col]).value for col in SRC}
        vals['Rule for LLM']=final[kid]
        vals['Deviation Level']=level[kid]
        for c,col in enumerate(OUT,1):
            cell=o.cell(ro,c,vals[col])
            if col in ('Rule for LLM','Protocol Reference & Quote','Description'):
                cell.alignment=Alignment(wrap_text=True,vertical='top')
        ro+=1
    o.column_dimensions['D'].width=38; o.column_dimensions['E'].width=66; o.column_dimensions['F'].width=44
outpath=f"{RUN}/ENX-CL-05-001_binary_distilled.xlsx"; wbo.save(outpath)

# drop audit
aud=Workbook(); a=aud.active; a.title='dropped'
for c,col in enumerate(['KRI ID','Sheet','Dropped stage','Reason'],1): a.cell(1,c,col).font=Font(bold=True)
ri=2
for r in kd:
    if not r['keep']:
        a.cell(ri,1,r['kri_id']); a.cell(ri,2,r['sheet']); a.cell(ri,3,r.get('dropped_stage','')); a.cell(ri,4,r.get('reason','')); ri+=1
a.column_dimensions['D'].width=90; aud.save(f"{RUN}/dropped_rules_audit.xlsx")

# changelog
cl=Workbook(); c2=cl.active; c2.title='changelog'
for c,col in enumerate(['KRI ID','Change','Reason'],1): c2.cell(1,c,col).font=Font(bold=True)
ri=2
c2.cell(ri,1,'ALL (298)'); c2.cell(ri,2,'footnote numbers stripped from Rule for LLM'); c2.cell(ri,3,'user: keep footnote content, omit footnote number'); ri+=1
for kid,ch,why in changelog:
    c2.cell(ri,1,kid); c2.cell(ri,2,ch); c2.cell(ri,3,why); ri+=1
c2.column_dimensions['B'].width=45; c2.column_dimensions['C'].width=40; cl.save(f"{RUN}/stage4_changelog.xlsx")

print("FINAL xlsx:", outpath)
print("rows per sheet:", {s.title:s.max_row-1 for s in wbo.worksheets})
print("audit + changelog written.")
