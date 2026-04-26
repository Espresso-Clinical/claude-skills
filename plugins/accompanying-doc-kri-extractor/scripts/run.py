#!/usr/bin/env python3
"""
Orchestration entry point for accompanying-doc-kri-extractor.

This script handles the DETERMINISTIC portions of the pipeline:
  - Stage 1 setup (output dir, run_config.json)
  - Stage 2 prep (extract-text: robust page-by-page PDF text)
  - Stage 3 mechanical tiering
  - Stage 5 SEMANTIC dedup (vs protocol golden set, then intra-document) — Gemini-based, jaccard is only a pre-filter
  - Stage 6 NDEF sweep (rule-based pre-screen)
  - Stage 7 C2 deterministic verbatim verification (HARD GATE — no LLM judging on C2)
  - Stage 7.5 minor-severity user review checkpoint
  - Stage 8 assembly (JSON + Excel)

The LLM-judging stages (5 Claude + 5 Gemini extraction, T2/T3/orphan/NDEF/verify
panels for C1/C3/C4/C5) are orchestrated by the top-level Claude instance per
SKILL.md — this script does not spawn them. C2 verbatim is REMOVED from LLM
panel duty and enforced deterministically here.

Subcommands:
  setup                 — Stage 1
  extract-text          — Stage 2 prep: robust per-page text extraction
  tier                  — Stage 3 mechanical tiering
  dedup-protocol        — Stage 5a (semantic, LLM)
  dedup-intra           — Stage 5b (semantic, LLM)
  ndef                  — Stage 6 rule-based pre-screen
  verify-c2             — Stage 7 deterministic verbatim gate (drops/auto-fixes pages)
  minor-review          — Stage 7.5 surface minor KRIs for user decision
  assemble              — Stage 8 final output
"""
import argparse, hashlib, json, os, sys, datetime, pathlib, re, unicodedata, warnings
warnings.filterwarnings("ignore")

DOC_TYPES = {
    "CMP": "Clinical Monitoring Plan",
    "CSMP": "Clinical Study Management Plan",
    "IMP": "IMP Handling Manual",
    "PDHP": "Protocol Deviation Handling Plan",
    "PV_PLAN": "Pharmacovigilance Plan",
    "SAP": "Statistical Analysis Plan",
    "PD_CLASS": "Protocol Deviation Classification Guide",
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def sha256_of_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def load_json(path: str):
    with open(path) as f:
        return json.load(f)


def dump_json(obj, path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False)


def normalize_text(s: str) -> str:
    """Normalize for similarity: lowercase, collapse whitespace, strip punctuation noise."""
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"[\u2018\u2019\u201C\u201D]", "'", s)    # smart quotes → straight
    s = re.sub(r"[\u2013\u2014\u2015]", "-", s)          # all dashes → hyphen
    s = re.sub(r"\s+", " ", s).strip()
    return s


def jaccard_tokens(a: str, b: str) -> float:
    ta = set(normalize_text(a).split())
    tb = set(normalize_text(b).split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def quote_verbatim_match(a: str, b: str) -> bool:
    """True if a's quote is present verbatim (after normalization) inside b's quote, or vice-versa."""
    na, nb = normalize_text(a), normalize_text(b)
    if not na or not nb:
        return False
    return na in nb or nb in na


def kris_equivalent(k1: dict, k2: dict, jaccard_threshold: float = 0.7) -> bool:
    """Two KRIs are treated as equivalent if either their rules overlap heavily OR quotes match."""
    if quote_verbatim_match(k1.get("supporting_quote", ""), k2.get("supporting_quote", "")):
        return True
    if jaccard_tokens(k1.get("rule_for_llm", ""), k2.get("rule_for_llm", "")) >= jaccard_threshold:
        return True
    return False


# ── Robust text + verbatim helpers (shared by extract-text and verify-c2) ────

def normalize_loose(s: str) -> str:
    """Whitespace-and-punctuation-stripped lowercase.

    This is what verify-c2 uses to match a quote against page text. Tolerates
    PDFs that lose spacing at line breaks (common with docx → PDF exports), and
    minor punctuation differences. If a quote string passes this test, the
    underlying obligation IS in the document — even if the original PDF text
    extraction collapsed spaces.
    """
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", s)
    rep = {"\u2018": "'", "\u2019": "'", "\u201C": '"', "\u201D": '"',
           "\u2013": "-", "\u2014": "-", "\u2212": "-", "\u00A0": " ",
           "\u2026": "...", "\u2009": " ", "\u200A": " ", "\u2002": " ", "\u2003": " "}
    for k, v in rep.items():
        s = s.replace(k, v)
    s = re.sub(r"[\s\W_]+", "", s).lower()
    return s


def quote_in_page(quote: str, page_loose: str) -> bool:
    """True if quote (after loose normalization) is substring of page (loose-normalized).

    Tries the full quote, then trims 1–2 words from each end (handles agents that
    over-grab leading/trailing words), then a 6-word head match (handles agents
    that cleanly compress mid-quote)."""
    qn = normalize_loose(quote)
    if not qn or len(qn) < 12:
        return False
    if qn in page_loose:
        return True
    # try trimming 1 then 2 words/segments from each end
    raw_words = re.sub(r"\s+", " ", quote.strip()).split()
    if len(raw_words) >= 8:
        for cut in (1, 2):
            t = normalize_loose(" ".join(raw_words[cut:-cut]))
            if t and t in page_loose:
                return True
    # head match: first ~30 chars of normalized quote
    head = qn[:30] if len(qn) > 30 else qn
    if head and len(head) >= 18 and head in page_loose:
        return True
    return False


def find_quote_pages(quote: str, pages_loose: dict) -> list:
    """Return list of page numbers where quote appears (loose match)."""
    return [p for p, t in pages_loose.items() if quote_in_page(quote, t)]


def extract_pdf_text_robust(pdf_path: str) -> dict:
    """Extract per-page text from a PDF using pdfplumber, with x_tolerance tuned
    for docx-converted PDFs (which often lose between-word spaces).

    Returns: {page_num: text_string}
    """
    try:
        import pdfplumber
    except ImportError:
        sys.exit("pdfplumber required. pip install pdfplumber")
    pages = {}
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            # x_tolerance=2 helps with tightly-kerned text; fallback raw extract if empty.
            try:
                t = page.extract_text(x_tolerance=2) or ""
            except Exception:
                t = page.extract_text() or ""
            pages[i] = t
    return pages


# ── Gemini call helper (shared by semantic dedup) ────────────────────────────

_GEMINI_SECRETS = os.path.expanduser("~/.claude/secrets/protocol-kri-extractor.json")


def _load_gemini():
    if not os.path.exists(_GEMINI_SECRETS):
        sys.exit(f"Gemini secrets missing: {_GEMINI_SECRETS}")
    with open(_GEMINI_SECRETS) as f:
        s = json.load(f)
    g = s.get("gemini", {})
    key = g.get("api_key", "")
    if not key or "PASTE" in key:
        sys.exit("Gemini api_key not set in secrets file")
    model = g.get("model_judge") or g.get("model", "gemini-2.5-pro")
    return key, model


def call_gemini(prompt: str, system: str, temperature: float = 0.1) -> str:
    from google import genai
    from google.genai import types
    key, model = _load_gemini()
    client = genai.Client(api_key=key)
    cfg = types.GenerateContentConfig(
        temperature=temperature, max_output_tokens=65536, system_instruction=system
    )
    try:
        cfg.thinking_config = types.ThinkingConfig(thinking_budget=16384)
    except Exception:
        pass
    r = client.models.generate_content(model=model, contents=prompt, config=cfg)
    parts = []
    try:
        for p in r.candidates[0].content.parts:
            if getattr(p, "thought", False):
                continue
            t = getattr(p, "text", None)
            if t:
                parts.append(t)
    except Exception:
        pass
    return "\n".join(parts) if parts else (r.text or "")


def parse_json_array(text: str) -> list:
    if not text:
        return []
    f = re.findall(r"```(?:json)?\s*\n([\s\S]*?)```", text)
    blob = f[0] if f else text
    s, e = blob.find("["), blob.rfind("]")
    if s >= 0 and e > s:
        try:
            return json.loads(blob[s:e + 1])
        except Exception:
            pass
    return []


# ── Stage assertions (real gate, not LLM theater) ────────────────────────────

def require(path: str, what: str):
    if not os.path.exists(path):
        sys.exit(f"BLOCK: {what} missing — expected at {path}. "
                 f"Previous stage did not produce its required artifact.")


def require_files_count(dir_path: str, glob_pat: str, expected: int, what: str):
    files = sorted(pathlib.Path(dir_path).glob(glob_pat))
    if len(files) != expected:
        sys.exit(f"BLOCK: {what} — expected {expected} files matching {glob_pat} in {dir_path}, "
                 f"found {len(files)}.")


# ── Stage 1: Setup ───────────────────────────────────────────────────────────

def cmd_setup(args):
    if args.doc_type not in DOC_TYPES:
        sys.exit(f"ERROR: --doc-type must be one of: {', '.join(DOC_TYPES)}")
    if not os.path.exists(args.pdf):
        sys.exit(f"ERROR: PDF not found: {args.pdf}")
    if not os.path.exists(args.protocol_golden_set):
        sys.exit(f"ERROR: protocol golden set not found: {args.protocol_golden_set}")

    gs = load_json(args.protocol_golden_set)
    if "kris" not in gs:
        sys.exit("ERROR: protocol golden set JSON must have a 'kris' key")

    out_dir = os.path.expanduser(
        f"~/Downloads/extractor/{args.protocol_id}/{args.run_id}/accompanying/{args.doc_type}"
    )
    os.makedirs(os.path.join(out_dir, "raw_extractions"), exist_ok=True)

    config = {
        "doc_type": args.doc_type,
        "doc_type_label": DOC_TYPES[args.doc_type],
        "pdf_path": os.path.abspath(args.pdf),
        "pdf_sha256": sha256_of_file(args.pdf),
        "protocol_golden_set_path": os.path.abspath(args.protocol_golden_set),
        "protocol_golden_set_sha256": sha256_of_file(args.protocol_golden_set),
        "protocol_golden_set_n_kris": len(gs["kris"]),
        "protocol_id": args.protocol_id,
        "run_id": args.run_id,
        "output_dir": out_dir,
        "started_at": datetime.datetime.utcnow().isoformat() + "Z",
        "pipeline_version": "0.1.0",
    }
    dump_json(config, os.path.join(out_dir, "run_config.json"))
    print(f"[setup] output dir: {out_dir}")
    print(f"[setup] doc type: {args.doc_type} ({DOC_TYPES[args.doc_type]})")
    print(f"[setup] protocol golden set loaded: {len(gs['kris'])} KRIs")
    print(f"[setup] run_config.json written.")


# ── Stage 3: Mechanical consensus tiering ────────────────────────────────────

def cmd_tier(args):
    out_dir = args.output_dir
    raw_dir = os.path.join(out_dir, "raw_extractions")
    files = sorted(pathlib.Path(raw_dir).glob("*.json"))
    if len(files) != 10:
        sys.exit(f"ERROR: expected 10 raw extraction files, found {len(files)}")

    all_agent_kris = []  # list of (agent_id, kri)
    for p in files:
        data = load_json(str(p))
        agent_id = data.get("agent_id", p.stem)
        for k in data.get("kris", []):
            all_agent_kris.append((agent_id, k))

    # Cluster: for each KRI, find all other KRIs semantically equivalent.
    clusters = []
    used = [False] * len(all_agent_kris)
    for i, (aid_i, k_i) in enumerate(all_agent_kris):
        if used[i]:
            continue
        cluster = [(aid_i, k_i)]
        used[i] = True
        for j in range(i + 1, len(all_agent_kris)):
            if used[j]:
                continue
            aid_j, k_j = all_agent_kris[j]
            if kris_equivalent(k_i, k_j):
                cluster.append((aid_j, k_j))
                used[j] = True
        clusters.append(cluster)

    # Assign tier by unique-agent count in the cluster
    tiered = []
    for idx, cluster in enumerate(clusters):
        unique_agents = {aid for aid, _ in cluster}
        n = len(unique_agents)
        if n >= 7:
            tier = "T1"
        elif n >= 4:
            tier = "T2"
        else:
            tier = "T3"
        # representative KRI: the one whose supporting_quote is longest (proxy for most specific)
        rep = max(cluster, key=lambda x: len(x[1].get("supporting_quote", "") or ""))[1]
        tiered.append({
            "cluster_idx": idx,
            "tier": tier,
            "n_agents": n,
            "agents": sorted(unique_agents),
            "representative_kri": rep,
            "all_variants": [k for _, k in cluster],
        })

    summary = {
        "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
        "n_raw_inputs": len(all_agent_kris),
        "n_clusters": len(clusters),
        "tier_counts": {
            "T1": sum(1 for t in tiered if t["tier"] == "T1"),
            "T2": sum(1 for t in tiered if t["tier"] == "T2"),
            "T3": sum(1 for t in tiered if t["tier"] == "T3"),
        },
        "clusters": tiered,
    }
    dump_json(summary, os.path.join(out_dir, "consensus_report.json"))
    print(f"[tier] wrote consensus_report.json — {summary['tier_counts']}")


# ── Stage 5a: SEMANTIC dedup vs protocol golden set (Gemini-backed) ──────────

def cmd_dedup_protocol(args):
    require(os.path.join(args.output_dir, "run_config.json"), "run_config.json (Stage 1)")
    require(args.candidates, "candidates JSON")
    cfg = load_json(os.path.join(args.output_dir, "run_config.json"))
    candidates = load_json(args.candidates)    # {"kris": [...]}
    protocol = load_json(cfg["protocol_golden_set_path"])

    # Phase 1 — fast deterministic pre-filter (verbatim quote match OR jaccard ≥0.85).
    # These are guaranteed dupes; semantic LLM pass handles the rest.
    pre_dropped = {}      # kri index → match record
    for i, k in enumerate(candidates.get("kris", [])):
        for p in protocol["kris"]:
            if quote_verbatim_match(k.get("supporting_quote", ""), p.get("supporting_quote", "")):
                pre_dropped[i] = {"matched_protocol_kri_id": p.get("kri_id"),
                                  "match_type": "quote_verbatim", "score": 1.0}
                break
            j = jaccard_tokens(k.get("rule_for_llm", ""), p.get("rule_for_llm", ""))
            if j >= 0.85:
                pre_dropped[i] = {"matched_protocol_kri_id": p.get("kri_id"),
                                  "match_type": "rule_jaccard", "score": j}
                break

    survivors_idx = [i for i in range(len(candidates.get("kris", []))) if i not in pre_dropped]
    survivors = [candidates["kris"][i] for i in survivors_idx]

    # Phase 2 — semantic dedup via Gemini on survivors (one batch call).
    # The LLM compares each accompanying-doc KRI against ALL protocol KRIs and
    # flags semantic equivalents that lexical matching missed (different
    # vocabulary, same obligation).
    sem_drops = {}       # kri_id-or-index in survivors → match record
    if survivors:
        acc_items = [{"id": k.get("kri_id") or f"_idx_{i}",
                      "name": k.get("kri_name"),
                      "rule": k.get("rule_for_llm")}
                     for i, k in enumerate(survivors)]
        prot_items = [{"id": p.get("kri_id"), "name": p.get("kri_name"),
                       "rule": p.get("rule_for_llm")} for p in protocol["kris"]]
        system = ("You are a clinical-trial KRI semantic-dedup judge. Compare "
                  "ACCOMPANYING-DOC KRIs to PROTOCOL GOLDEN SET KRIs and flag any accompanying "
                  "KRI that is semantically equivalent to a protocol KRI (same underlying "
                  "monitoring check, even if worded differently or scoped slightly differently). "
                  "Be CONSERVATIVE — same TOPIC but different OBLIGATIONS is NOT a dupe (e.g., "
                  "storage temperature vs alarm settings are different obligations). Output JSON only.")
        prompt = ("ACCOMPANYING KRIS:\n" + json.dumps(acc_items, indent=1) +
                  "\n\nPROTOCOL GOLDEN SET KRIS:\n" + json.dumps(prot_items, indent=1) +
                  "\n\nReturn a JSON array. Each element: "
                  "{\"acc_id\": \"<accompanying_id>\", \"protocol_id\": \"<protocol_id>\", "
                  "\"reason\": \"<why they overlap>\"}. Output JSON array only.")
        try:
            text = call_gemini(prompt, system, temperature=0.1)
            for rec in parse_json_array(text):
                aid = rec.get("acc_id")
                # find which survivor this is
                for k in survivors:
                    if (k.get("kri_id") or "") == aid:
                        sem_drops[id(k)] = {"matched_protocol_kri_id": rec.get("protocol_id"),
                                            "match_type": "semantic_llm",
                                            "reason": rec.get("reason", "")}
                        break
        except Exception as e:
            print(f"[dedup-protocol] WARNING: semantic pass failed ({e}); using lexical only.")

    dropped = []
    kept = []
    for i, k in enumerate(candidates.get("kris", [])):
        if i in pre_dropped:
            r = pre_dropped[i]
            dropped.append({"candidate": k, **r})
        elif id(k) in sem_drops:
            r = sem_drops[id(k)]
            dropped.append({"candidate": k, **r})
        else:
            kept.append(k)

    report = {
        "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
        "n_candidates_in": len(candidates.get("kris", [])),
        "n_dropped": len(dropped),
        "n_dropped_lexical": sum(1 for d in dropped if d["match_type"] != "semantic_llm"),
        "n_dropped_semantic": sum(1 for d in dropped if d["match_type"] == "semantic_llm"),
        "n_kept": len(kept),
        "dropped": dropped,
    }
    dump_json(report, os.path.join(args.output_dir, "protocol_dedup_report.json"))
    dump_json({"kris": kept}, os.path.join(args.output_dir, "after_protocol_dedup.json"))
    print(f"[dedup-protocol] dropped {len(dropped)} ({report['n_dropped_lexical']} lexical + {report['n_dropped_semantic']} semantic), kept {len(kept)}")


# ── Stage 5b: SEMANTIC intra-document dedup (Gemini-backed) ──────────────────

def cmd_dedup_intra(args):
    require(os.path.join(args.output_dir, "after_protocol_dedup.json"),
            "after_protocol_dedup.json (Stage 5a must run before 5b)")
    data = load_json(os.path.join(args.output_dir, "after_protocol_dedup.json"))
    kris = data["kris"]

    # Phase 1 — fast lexical pre-cluster (jaccard + verbatim quote match) collapses obvious dupes.
    used = [False] * len(kris)
    pre_clusters = []
    for i, k_i in enumerate(kris):
        if used[i]:
            continue
        cluster = [i]
        used[i] = True
        for j in range(i + 1, len(kris)):
            if used[j]:
                continue
            if kris_equivalent(k_i, kris[j]):
                cluster.append(j)
                used[j] = True
        pre_clusters.append(cluster)
    # Survivors after lexical pre-cluster: keep longest from each cluster.
    pre_survivors = []
    pre_merges = []
    for cl in pre_clusters:
        s_idx = max(cl, key=lambda x: (len(kris[x].get("rule_for_llm", "") or ""),
                                       len(kris[x].get("supporting_quote", "") or "")))
        pre_survivors.append(kris[s_idx])
        if len(cl) > 1:
            pre_merges.append({"kept_kri_name": kris[s_idx].get("kri_name"),
                               "merged_kri_names": [kris[x].get("kri_name") for x in cl if x != s_idx]})

    # Phase 2 — semantic dedup via Gemini on lexical-survivors. The LLM finds
    # clusters of semantically equivalent KRIs that lexical matching missed
    # (synonyms / different word order / scope variants).
    semantic_drops = set()
    semantic_clusters = []
    if pre_survivors:
        items = [{"id": k.get("kri_id") or f"_idx_{i}", "name": k.get("kri_name"),
                  "rule": k.get("rule_for_llm")} for i, k in enumerate(pre_survivors)]
        system = ("You are a clinical-trial KRI semantic-dedup judge. Find clusters of "
                  "semantically equivalent KRIs — same monitorable obligation, even if worded "
                  "differently. Two rules are DUPLICATES if a CRA verifying one would be "
                  "verifying the same underlying check. Different KRIs covering the SAME clause "
                  "from different angles are duplicates. Be CONSERVATIVE — distinct atomic "
                  "obligations within one process are NOT duplicates. Output JSON only.")
        prompt = ("INPUT KRIS:\n" + json.dumps(items, indent=1) +
                  "\n\nFind ALL clusters of semantic duplicates. Output a JSON array. Each "
                  "element: {\"primary\": \"<id_to_keep>\", \"duplicates\": [\"<id_to_drop>\", ...], "
                  "\"reason\": \"<why same check>\"}. Only include clusters with ≥2 KRIs. "
                  "Output JSON array only.")
        try:
            text = call_gemini(prompt, system, temperature=0.1)
            id_to_idx = {k.get("kri_id"): i for i, k in enumerate(pre_survivors)}
            for cl in parse_json_array(text):
                primary = cl.get("primary")
                dups = cl.get("duplicates", [])
                if primary in id_to_idx:
                    semantic_clusters.append(cl)
                    for d in dups:
                        if d in id_to_idx and d != primary:
                            semantic_drops.add(d)
        except Exception as e:
            print(f"[dedup-intra] WARNING: semantic pass failed ({e}); lexical-only output.")

    final = [k for k in pre_survivors if (k.get("kri_id") or "") not in semantic_drops]

    report = {
        "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
        "n_in": len(kris),
        "n_after_lexical": len(pre_survivors),
        "n_out": len(final),
        "lexical_merges": pre_merges,
        "semantic_clusters": semantic_clusters,
        "semantic_dropped_ids": sorted(semantic_drops),
    }
    dump_json(report, os.path.join(args.output_dir, "intra_dedup_report.json"))
    dump_json({"kris": final}, os.path.join(args.output_dir, "after_intra_dedup.json"))
    print(f"[dedup-intra] in={len(kris)}, lexical→{len(pre_survivors)}, semantic→{len(final)} "
          f"({len(pre_merges)} lexical merges + {len(semantic_clusters)} semantic clusters)")


# ── Stage 6: NDEF candidate pre-screen (regex Pass 6.1 only) ─────────────────
#
# This script implements ONLY Pass 6.1 (regex candidate flagging). The
# definitive classification is done by the 6-judge panel in Pass 6.2 — the
# top-level Claude orchestrator is responsible for that, per SKILL.md Stage 6.
# This script's output is consumed by the panel as the candidate set.
#
# Design principles:
#   - Be liberal on candidate flagging: false-positive candidates are filtered
#     by the panel; missed candidates risk leaving NDEF KRIs in the main list.
#   - Default disposition for non-candidates remains DEFINABLE — the panel
#     samples 10% of non-candidates as a false-negative check.
#   - Triggers cover the protocol skill's NDEF criteria AND user-named
#     classes ("at the discretion of doctor/pharmacist/CRA", "according to
#     clinical need", "best efforts", etc.).
#   - Borderline phrases that often appear next to definable thresholds
#     ("as required", "as necessary" alone) are intentionally NOT triggers —
#     the panel handles them.

NDEF_TRIGGERS = [
    # Discretion language — clear NDEF
    r"\bat (?:the )?(?:sole )?discretion of\b",
    r"\bat the (?:sponsor'?s|investigator'?s|cra'?s|pharmacist'?s|doctor'?s|physician'?s) discretion\b",
    r"\bas deemed (?:appropriate|necessary|fit|required)\b",
    r"\bif deemed (?:appropriate|necessary|fit|required)\b",
    # Investigator clinical judgment
    r"\bin the (?:investigator'?s|physician'?s|doctor'?s|pi'?s) (?:opinion|judgment|judgement)\b",
    r"\bif clinically (?:significant|indicated|relevant|warranted)\b",
    r"\bclinically (?:significant|relevant|indicated)\b",
    r"\bper clinical judg(?:e)?ment\b",
    r"\bclinical(?:ly)? judg(?:e)?ment\b",
    # Undefined time windows
    r"\bas soon as possible\b",
    r"\bin a timely manner\b",
    r"\bpromptly\b",
    r"\bwithout undue delay\b",
    r"\bwithin a reasonable (?:time|period)\b",
    r"\bin due (?:time|course)\b",
    # Undefined effort / quantity
    r"\b(?:reasonable|best|great) effort(?:s)?\b",
    r"\bevery effort\b",
    r"\b(?:adequate|sufficient)(?:ly)?\b",  # standalone — panel filters when modifier
    r"\bappropriate(?:ly)?\b(?!\s+(?:section|table|level\s*\d|procedure|form))",  # standalone
    # "According to / as needed" — when standalone with no concrete trigger
    r"\baccording to clinical need\b",
    r"\baccording to (?:clinical|medical) (?:judg(?:e)?ment|practice)\b",
    r"\bas (?:clinically )?(?:needed|required|indicated)\b",
    r"\bif (?:medically|clinically) (?:needed|required|indicated)\b",
    # "Where feasible / possible / applicable" without a concrete bound
    r"\bwhere (?:feasible|possible|applicable|practical|practicable)\b",
    r"\bwhenever (?:feasible|possible|practical|practicable)\b",
    # Subjective / soft thresholds
    r"\bsubject to\b.*\b(?:judg(?:e)?ment|review|approval|interpretation)\b",
    r"\bshould (?:consider|use (?:their )?judg(?:e)?ment|exercise judg(?:e)?ment)\b",
    r"\bwhen (?:appropriate|necessary|required)\b.*\bjudg(?:e)?ment\b",
    # Per-local-practice (often NDEF; panel decides)
    r"\bper local (?:practice|standard|guidance|sop)\b",
    r"\bper (?:institutional|local) policy\b",
]
NDEF_RES = [re.compile(p, re.IGNORECASE) for p in NDEF_TRIGGERS]


def cmd_ndef(args):
    """Pass 6.1 — regex candidate pre-screen.

    Marks each KRI with `ndef_candidate` (bool) and matched `trigger_phrases`.
    Does NOT set `ndef` (final classification) — that is done by the 6-judge
    panel (Pass 6.2) orchestrated by Claude per SKILL.md Stage 6.
    """
    data = load_json(os.path.join(args.output_dir, "after_intra_dedup.json"))
    kris = data["kris"]

    results = []
    for k in kris:
        combined = " ".join([k.get("rule_for_llm", "") or "", k.get("supporting_quote", "") or ""])
        hits = []
        for pat in NDEF_RES:
            m = pat.search(combined)
            if m:
                hits.append(m.group(0))
        candidate = bool(hits)
        # Pass 6.1 sets candidate flag and trigger phrases ONLY.
        # Final ndef classification is set by Pass 6.2 panel.
        k["ndef_candidate"] = candidate
        k["ndef_trigger_phrases"] = hits
        # ndef remains False until the panel decides.
        if "ndef" not in k:
            k["ndef"] = False
        results.append({
            "kri_name": k.get("kri_name"),
            "ndef_candidate": candidate,
            "trigger_phrases": hits,
        })

    report = {
        "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
        "pass": "6.1 — regex candidate pre-screen",
        "n_kris": len(kris),
        "n_candidates": sum(1 for r in results if r["ndef_candidate"]),
        "note": "Final NDEF classification is set by the Pass 6.2 6-judge panel (orchestrated by Claude per SKILL.md Stage 6). This file is the candidate input for that panel.",
        "per_kri": results,
    }
    dump_json(report, os.path.join(args.output_dir, "ndef_sweep_report.json"))
    dump_json({"kris": kris}, os.path.join(args.output_dir, "after_ndef.json"))
    print(f"[ndef-prescreen] total={len(kris)}, candidates={report['n_candidates']}")
    print(f"[ndef-prescreen] Pass 6.2 (6-judge panel) must run before Stage 7.")


# ── Stage 8: Assemble ────────────────────────────────────────────────────────

def cmd_assemble(args):
    cfg = load_json(os.path.join(args.output_dir, "run_config.json"))
    # Prefer the most-recent post-gate file in this priority:
    #   after_minor_review.json (Stage 7.5)
    #   > after_c2.json         (Stage 7 deterministic verbatim gate)
    #   > after_ndef.json       (Stage 6 — legacy fallback)
    for candidate in ("after_minor_review.json", "after_c2.json", "after_ndef.json"):
        p = os.path.join(args.output_dir, candidate)
        if os.path.exists(p):
            data = load_json(p)
            print(f"[assemble] reading from {candidate}")
            break
    else:
        sys.exit("BLOCK: no post-pipeline artifact found (need after_minor_review/after_c2/after_ndef).")
    kris = data["kris"]

    main_kris = [k for k in kris if not k.get("ndef", False)]
    ndef_kris = [k for k in kris if k.get("ndef", False)]

    # Assign IDs in combined order (main first, then ndef)
    doc_type = cfg["doc_type"]
    seq = 1
    for group in (main_kris, ndef_kris):
        for k in group:
            k["kri_id"] = f"{doc_type}-{seq:03d}"
            # Recompute combined_ref defensively
            dr = (k.get("document_reference") or "").strip()
            sq = (k.get("supporting_quote") or "").strip().strip('"')
            k["supporting_quote"] = sq
            k["combined_ref"] = f'{dr} — "{sq}"' if dr and sq else ""
            k["doc_type"] = doc_type
            k["doc_type_label"] = cfg["doc_type_label"]
            seq += 1

    out = {
        "meta": {
            "doc_type": doc_type,
            "doc_type_label": cfg["doc_type_label"],
            "protocol_id": cfg["protocol_id"],
            "run_id": cfg["run_id"],
            "protocol_golden_set_sha256": cfg["protocol_golden_set_sha256"],
            "pdf_sha256": cfg["pdf_sha256"],
            "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
            "n_kris": len(main_kris),
            "n_ndef_kris": len(ndef_kris),
        },
        "kris": main_kris,
        "ndef_kris": ndef_kris,
    }
    dump_json(out, os.path.join(args.output_dir, "accompanying_golden_set.json"))

    # Excel — column structure mirrors the protocol-kri-extractor skill:
    #   KRI ID | Category | KRI Name | Description | Rule for LLM | Document Reference & Quote
    # NDEF KRIs appear as a second table on the SAME sheet, below the main
    # table (separated by a blank row + a "Non-Definable KRIs" header row).
    # Severity is preserved in JSON only — it is NOT a column in the Excel,
    # matching the protocol skill's column set.
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "KRIs"

        HEADER = ["KRI ID", "Category", "KRI Name", "Description",
                  "Rule for LLM", "Document Reference & Quote"]

        def kri_row(k):
            return [k.get("kri_id"), k.get("doc_type_label"), k.get("kri_name"),
                    k.get("description"), k.get("rule_for_llm"), k.get("combined_ref")]

        bold = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ndef_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        wrap = Alignment(wrap_text=True, vertical="top")

        # ── Main table ────────────────────────────────────────────────────────
        ws.append(HEADER)
        for c in range(1, len(HEADER) + 1):
            ws.cell(1, c).font = bold
            ws.cell(1, c).fill = fill
            ws.cell(1, c).alignment = Alignment(horizontal="center", wrap_text=True)
        for k in main_kris:
            ws.append(kri_row(k))

        # ── Spacer + NDEF sub-table on same sheet ────────────────────────────
        if ndef_kris:
            ws.append([])  # blank row
            label_row = ws.max_row + 1
            ws.cell(label_row, 1, "Non-Definable KRIs").font = Font(bold=True, color="FFFFFF", size=12)
            ws.cell(label_row, 1).fill = ndef_fill
            ws.merge_cells(start_row=label_row, start_column=1,
                           end_row=label_row, end_column=len(HEADER))
            ws.append(HEADER)
            sub_header_row = ws.max_row
            for c in range(1, len(HEADER) + 1):
                ws.cell(sub_header_row, c).font = bold
                ws.cell(sub_header_row, c).fill = ndef_fill
                ws.cell(sub_header_row, c).alignment = Alignment(horizontal="center", wrap_text=True)
            for k in ndef_kris:
                ws.append(kri_row(k))

        # Apply wrap to all data rows + reasonable column widths
        for col_letter, width in zip("ABCDEF", [12, 22, 28, 60, 60, 50]):
            ws.column_dimensions[col_letter].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap

        # ── Dropped-vs-protocol sheet (audit trail of Stage 5a) ──────────────
        prot_dedup_path = os.path.join(args.output_dir, "protocol_dedup_report.json")
        if os.path.exists(prot_dedup_path):
            rep = load_json(prot_dedup_path)
            ws3 = wb.create_sheet("Dropped (vs protocol)")
            ws3.append(["Candidate KRI Name", "Matched Protocol KRI ID", "Match Type", "Score",
                        "Candidate Rule", "Candidate Quote"])
            for c in range(1, 7):
                ws3.cell(1, c).font = bold
                ws3.cell(1, c).fill = fill
            for row in rep.get("dropped", []):
                c = row["candidate"]
                ws3.append([c.get("kri_name"), row.get("matched_protocol_kri_id"),
                            row.get("match_type"), row.get("score"),
                            c.get("rule_for_llm"), c.get("supporting_quote")])

        xlsx_path = os.path.join(args.output_dir, "Accompanying_KRIs.xlsx")
        wb.save(xlsx_path)
        print(f"[assemble] Excel: {xlsx_path}")
    except ImportError:
        print("[assemble] WARNING: openpyxl not installed; skipping Excel output.")

    print(f"[assemble] golden set: {len(main_kris)} KRIs + {len(ndef_kris)} NDEF")
    print(f"[assemble] output dir: {args.output_dir}")


# ── Stage 2 prep: robust per-page text extraction ────────────────────────────

def cmd_extract_text(args):
    require(os.path.join(args.output_dir, "run_config.json"), "run_config.json (Stage 1)")
    cfg = load_json(os.path.join(args.output_dir, "run_config.json"))
    pages = extract_pdf_text_robust(cfg["pdf_path"])
    out_path = os.path.join(args.output_dir, "_pdf_text.txt")
    with open(out_path, "w") as f:
        for i in sorted(pages):
            f.write(f"\n===== PAGE {i} =====\n{pages[i]}\n")
    # Also save loose-normalized index for verify-c2 reuse.
    pages_loose = {p: normalize_loose(t) for p, t in pages.items()}
    dump_json({"pages_loose": pages_loose}, os.path.join(args.output_dir, "_pdf_pages_loose.json"))
    print(f"[extract-text] wrote {out_path} ({len(pages)} pages, {sum(len(t) for t in pages.values())} chars)")


# ── Stage 7 (C2): DETERMINISTIC verbatim verification ────────────────────────
#
# This is the gate that previously failed when implemented as an LLM panel.
# C1 (reference correctness), C3 (rule correctness), C4 (description fidelity),
# and C5 (schema) are still LLM-judged by the panel that the orchestrator runs;
# C2 is enforced HERE, in code, with a substring check. No exceptions.

def cmd_verify_c2(args):
    require(os.path.join(args.output_dir, "run_config.json"), "run_config.json (Stage 1)")
    require(args.input, "verify input JSON (e.g. after_intra_dedup.json or after_ndef.json)")
    # If pages_loose cache is missing, regenerate it on the fly.
    pl_path = os.path.join(args.output_dir, "_pdf_pages_loose.json")
    if not os.path.exists(pl_path):
        cfg = load_json(os.path.join(args.output_dir, "run_config.json"))
        pages = extract_pdf_text_robust(cfg["pdf_path"])
        pages_loose = {p: normalize_loose(t) for p, t in pages.items()}
        dump_json({"pages_loose": pages_loose}, pl_path)
    pages_loose = {int(k): v for k, v in load_json(pl_path)["pages_loose"].items()}

    data = load_json(args.input)
    kris = data.get("kris", [])

    kept = []
    dropped = []
    fixed_pages = []
    for k in kris:
        q = k.get("supporting_quote", "")
        found_pages = find_quote_pages(q, pages_loose)
        if not found_pages:
            dropped.append({"kri_name": k.get("kri_name"),
                            "kri_id": k.get("kri_id"),
                            "quote_excerpt": (q or "")[:80],
                            "cited_ref": k.get("document_reference"),
                            "reason": "supporting_quote not found anywhere in document"})
            continue
        # Auto-fix wrong page citations.
        ref = k.get("document_reference") or ""
        m = re.search(r"p\.?\s*(\d+)", ref, re.I)
        cited = int(m.group(1)) if m else None
        if cited is None or cited not in found_pages:
            actual = found_pages[0]
            new_ref = re.sub(r"(p\.?\s*)\d+", f"p.{actual}", ref, count=1, flags=re.I) if cited else f"{ref} (p.{actual})"
            fixed_pages.append({"kri_name": k.get("kri_name"),
                                "old_ref": ref, "new_ref": new_ref,
                                "cited_page": cited, "actual_pages": found_pages})
            k["document_reference"] = new_ref
            sq = (k.get("supporting_quote") or "").strip().strip('"')
            k["combined_ref"] = f'{new_ref} — "{sq}"' if sq else ""
        kept.append(k)

    report = {
        "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
        "n_in": len(kris),
        "n_kept": len(kept),
        "n_dropped": len(dropped),
        "n_pages_auto_fixed": len(fixed_pages),
        "dropped": dropped,
        "page_fixes": fixed_pages,
    }
    dump_json(report, os.path.join(args.output_dir, "verify_c2_report.json"))
    dump_json({"kris": kept}, os.path.join(args.output_dir, "after_c2.json"))
    print(f"[verify-c2] in={len(kris)}, kept={len(kept)}, dropped={len(dropped)}, page_fixes={len(fixed_pages)}")


# ── Stage 7.5: minor-severity user review checkpoint ─────────────────────────
#
# Surfaces every minor-severity KRI for orchestrator → user decision. The
# orchestrator (Claude) is responsible for presenting the list and recording
# the user's keep/drop choices in `minor_review_decisions.json`. If that
# decisions file is present and complete, this command applies it; otherwise
# it produces the prompt artifact and exits non-fatally so the orchestrator
# can ask the user.

def cmd_minor_review(args):
    require(os.path.join(args.output_dir, "after_c2.json"),
            "after_c2.json (Stage 7 verify-c2 must run first)")
    data = load_json(os.path.join(args.output_dir, "after_c2.json"))
    kris = data["kris"]
    minors = [k for k in kris if (k.get("severity") or "").lower() == "minor"]

    decisions_path = os.path.join(args.output_dir, "minor_review_decisions.json")
    if os.path.exists(decisions_path):
        decisions = load_json(decisions_path)
        keep_set = set(decisions.get("keep", []))
        # apply: drop any minor KRI not explicitly kept.
        kept = []
        dropped = []
        for k in kris:
            sev = (k.get("severity") or "").lower()
            kid = k.get("kri_id") or k.get("kri_name")
            if sev == "minor" and kid not in keep_set:
                dropped.append({"kri_id": kid, "kri_name": k.get("kri_name")})
                continue
            kept.append(k)
        report = {"applied_decisions": True, "n_minor_total": len(minors),
                  "n_minor_kept": len(minors) - len(dropped),
                  "n_minor_dropped": len(dropped),
                  "dropped": dropped}
        dump_json(report, os.path.join(args.output_dir, "minor_review_report.json"))
        dump_json({"kris": kept}, os.path.join(args.output_dir, "after_minor_review.json"))
        print(f"[minor-review] applied user decisions: kept {len(minors)-len(dropped)}/{len(minors)} minor KRIs")
    else:
        # produce prompt artifact for the orchestrator
        prompt = {
            "n_minor_total": len(minors),
            "instructions": ("Review each minor-severity KRI below. For each, decide KEEP or DROP. "
                             "Then write minor_review_decisions.json to this directory with shape: "
                             "{\"keep\": [\"IMP-XYZ\", ...]} listing only the kri_ids to keep. "
                             "Re-run `run.py minor-review --output-dir <dir>` after writing decisions."),
            "minor_kris": [{"kri_id": k.get("kri_id"), "kri_name": k.get("kri_name"),
                            "description": k.get("description"),
                            "rule_for_llm": k.get("rule_for_llm")}
                           for k in minors],
        }
        dump_json(prompt, os.path.join(args.output_dir, "minor_review_prompt.json"))
        # Also write a no-op pass-through so assemble can proceed without the gate
        # if the orchestrator chooses (e.g., user declines to review).
        dump_json({"kris": kris}, os.path.join(args.output_dir, "after_minor_review.json"))
        print(f"[minor-review] {len(minors)} minor KRIs surfaced → minor_review_prompt.json")
        print(f"[minor-review] orchestrator must collect user decisions, write minor_review_decisions.json, then re-run this command")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Accompanying-doc KRI extractor orchestration")
    sub = ap.add_subparsers(dest="cmd", required=True)

    s = sub.add_parser("setup")
    s.add_argument("--pdf", required=True)
    s.add_argument("--doc-type", required=True)
    s.add_argument("--protocol-golden-set", required=True)
    s.add_argument("--protocol-id", required=True)
    s.add_argument("--run-id", required=True)
    s.set_defaults(func=cmd_setup)

    s = sub.add_parser("tier")
    s.add_argument("--output-dir", required=True)
    s.set_defaults(func=cmd_tier)

    s = sub.add_parser("dedup-protocol")
    s.add_argument("--output-dir", required=True)
    s.add_argument("--candidates", required=True, help="JSON with {'kris': [...]} of the post-tiering/orphan candidate set")
    s.set_defaults(func=cmd_dedup_protocol)

    s = sub.add_parser("dedup-intra")
    s.add_argument("--output-dir", required=True)
    s.set_defaults(func=cmd_dedup_intra)

    s = sub.add_parser("ndef")
    s.add_argument("--output-dir", required=True)
    s.set_defaults(func=cmd_ndef)

    s = sub.add_parser("extract-text")
    s.add_argument("--output-dir", required=True)
    s.set_defaults(func=cmd_extract_text)

    s = sub.add_parser("verify-c2")
    s.add_argument("--output-dir", required=True)
    s.add_argument("--input", required=True,
                   help="Path to KRIs JSON to verify (typically after_ndef.json)")
    s.set_defaults(func=cmd_verify_c2)

    s = sub.add_parser("minor-review")
    s.add_argument("--output-dir", required=True)
    s.set_defaults(func=cmd_minor_review)

    s = sub.add_parser("assemble")
    s.add_argument("--output-dir", required=True)
    s.set_defaults(func=cmd_assemble)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
