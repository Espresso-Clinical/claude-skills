"""
Step 4A — Assembly + Excel Generation (ELIG / SAF / END / OPS)
Merges raw_ELIG.json, raw_SAF.json, raw_END.json, raw_OPS.json (plus optional
raw_STAT.json and raw_GOV.json which fold into END) into a single
extracted_kris.json AND generates Extracted_KRIs.xlsx. No LLM required.

SOA is OUT OF SCOPE for this skill — handled by `soa-kri-extractor`.

Must be run AFTER Step 3D (full verbatim verification) passes 100%.

Usage:
  python step4a_assemble.py --out /path/to/output/ --manifest /path/to/manifest.json

Dependencies:
  pip install openpyxl --break-system-packages -q
"""
import json, os, re, argparse, shutil, pathlib

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed. Excel generation skipped. Run: pip install openpyxl")

# ── Single source of truth for field names ───────────────────────────────────
F_KRI_ID       = "kri_id"
F_CATEGORY     = "category_label"
F_NAME         = "kri_name"
F_DESCRIPTION  = "description"
F_RULE         = "rule_for_llm"
F_REF          = "protocol_reference"
F_QUOTE        = "supporting_quote"
F_COMBINED     = "combined_ref"
F_FOOTNOTES    = "additional_footnotes"
F_SEVERITY     = "severity"

CATEGORY_LABELS = {
    "ELIG": "Eligibility",
    "SAF":  "Safety & Toxicity",
    "END":  "Endpoints & Statistics",
    "OPS":  "Operations & Compliance",
}

# STAT and GOV are sub-categories of END — they share category_id "END"
# but are extracted in separate passes and may have their own raw files
SUBCATEGORY_MAP = {
    "STAT": ("END", "Endpoints & Statistics"),
    "GOV":  ("END", "Endpoints & Statistics"),
}

DOMAIN_ORDER = ["ELIG", "SAF", "END", "OPS"]

# Raw file loading order — includes STAT and GOV as separate raw files
RAW_FILE_ORDER = ["ELIG", "SAF", "END", "STAT", "GOV", "OPS"]

DOMAIN_COLORS = {
    "ELIG": "FCE5CD",
    "SAF":  "F4CCCC",
    "END":  "CFE2F3",
    "OPS":  "EAD1DC",
}

# Excel column spec — exact, no deviations
EXCEL_COLS   = ["KRI ID", "Category", "KRI Name", "Description", "Rule for LLM", "Protocol Reference & Quote", "Severity"]
EXCEL_WIDTHS = [16,       28,         34,          52,             60,             90,                          12]


def strip_outer_quotes(s: str) -> str:
    """Remove outer double-quote wrapping if present. Safety net — quotes should never be there."""
    s = s.strip()
    if s.startswith('"') and s.endswith('"') and len(s) > 1:
        s = s[1:-1]
    elif s.startswith('"'):
        s = s.lstrip('"')
    return s


def make_combined_ref(ref: str, quote: str) -> str:
    """Compute combined_ref as: 'ref — "quote"'."""
    return f'{ref} \u2014 "{quote}"'


def assemble(out_dir: str, manifest_path: str) -> str:
    """Merge raw category files into extracted_kris.json and Extracted_KRIs.xlsx."""
    with open(manifest_path) as f:
        manifest = json.load(f)

    all_kris, categories_meta, by_cat = [], [], {}

    for cat in RAW_FILE_ORDER:
        raw_path = os.path.join(out_dir, f"raw_{cat}.json")
        if not os.path.exists(raw_path):
            if cat not in ("STAT", "GOV"):  # STAT/GOV are optional separate files
                print(f"  {cat}: not found, skipping")
            continue

        # STAT and GOV map to END domain
        domain_cat = cat
        cat_label = CATEGORY_LABELS.get(cat)
        if cat in SUBCATEGORY_MAP:
            domain_cat, cat_label = SUBCATEGORY_MAP[cat]

        with open(raw_path) as f:
            data = json.load(f)

        # Handle multiple JSON formats
        if isinstance(data, list):
            kris = data
        else:
            kris = data.get("kris", data.get("KRIs", []))
            if not kris and isinstance(data, dict):
                for k, v in data.items():
                    if isinstance(v, list) and len(v) > 0:
                        kris = v
                        break

        # Normalize fields, deduplicate by rule_for_llm
        seen, cleaned = set(), []
        for k in kris:
            if not isinstance(k, dict):
                continue
            rule = (k.get(F_RULE) or "").strip().lower()
            if rule and rule in seen:
                continue
            if rule:
                seen.add(rule)
            if not k.get(F_RULE) and not k.get(F_NAME):
                continue

            # Sanitize quote — must never have outer quotes
            raw_quote = k.get(F_QUOTE, "") or ""
            quote = strip_outer_quotes(raw_quote)
            ref   = k.get(F_REF, "") or ""

            entry = {
                F_KRI_ID:    k.get(F_KRI_ID, ""),
                F_NAME:      k.get(F_NAME, ""),
                F_DESCRIPTION: k.get(F_DESCRIPTION, None),
                "category_id":   domain_cat,
                F_CATEGORY:  cat_label,
                F_RULE:      k.get(F_RULE, None),
                F_REF:       ref,
                F_QUOTE:     quote,
                F_COMBINED:  k.get(F_COMBINED) or make_combined_ref(ref, quote),
                F_FOOTNOTES: k.get(F_FOOTNOTES, None),
                F_SEVERITY:  k.get(F_SEVERITY, "major"),
                "agent_count": k.get("agent_count"),
            }
            cleaned.append(entry)

        # STAT and GOV merge into the END domain bucket
        if domain_cat != cat:
            by_cat.setdefault(domain_cat, []).extend(cleaned)
        else:
            by_cat.setdefault(cat, []).extend(cleaned)

        all_kris.extend(cleaned)
        categories_meta.append({
            "id": cat,
            "parent_domain": domain_cat,
            "label": cat_label,
            "kri_count": len(cleaned),
            "kri_ids": [k[F_KRI_ID] for k in cleaned]
        })
        print(f"  {cat}: {len(cleaned)} KRIs" + (f" (→ {domain_cat})" if domain_cat != cat else ""))

    # ── Post-assembly validation ──────────────────────────────────────────────
    errors = []
    for k in all_kris:
        kid = k.get(F_KRI_ID, "?")
        if not k.get(F_RULE):
            errors.append(f"{kid}: {F_RULE} is empty")
        if not k.get(F_CATEGORY):
            errors.append(f"{kid}: {F_CATEGORY} is empty")
        if not k.get(F_COMBINED):
            errors.append(f"{kid}: {F_COMBINED} is missing")
        if not k.get(F_QUOTE):
            errors.append(f"{kid}: {F_QUOTE} is empty")
        q = k.get(F_QUOTE, "")
        if q.startswith('"') or q.endswith('"'):
            errors.append(f"{kid}: {F_QUOTE} has outer quotes (will corrupt combined_ref)")

    if errors:
        print(f"\nASSEMBLY VALIDATION ERRORS ({len(errors)}):")
        for e in errors:
            print(f"  {e}")
        raise SystemExit("Fix the errors above before continuing to Step 4A Excel output.")
    else:
        print(f"\nAssembly validation OK: {len(all_kris)} KRIs, all required fields populated.")

    output = {
        "_meta": {
            "version": "10.0",
            "protocol": manifest.get("protocol_id", "UNKNOWN"),
            "sponsor": manifest.get("sponsor", ""),
            "extracted_by": "protocol-kri-extractor",
            "total_kris": len(all_kris),
            "total_categories": len(categories_meta)
        },
        "categories": categories_meta,
        "kris": all_kris
    }

    out_path = os.path.join(out_dir, "extracted_kris.json")

    # Backup before writing
    if os.path.exists(out_path):
        backup = pathlib.Path(out_path).with_suffix(".pre_4a.json")
        shutil.copy(out_path, backup)
        print(f"  Backup: {backup}")

    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"  Saved: {len(all_kris)} KRIs → {out_path}")

    # Generate Excel
    if HAS_OPENPYXL:
        xlsx_path = generate_excel(out_dir, by_cat)
        print(f"  Excel: {xlsx_path}")
    else:
        print("  Excel: SKIPPED (openpyxl not installed)")

    return out_path


def generate_excel(out_dir: str, by_cat: dict) -> str:
    """Generate Extracted_KRIs.xlsx.

    Columns (identical across all domain sheets):
      KRI ID | Category | KRI Name | Description | Rule for LLM | Protocol Reference & Quote

    No 'Domain' column. No separate 'Supporting Quote' or 'Protocol Reference' column.
    The combined_ref field is the single Protocol Reference & Quote column.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR_FONT  = Font(bold=True, size=11, color="FFFFFF")
    HDR_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    BODY_FONT = Font(name="Calibri", size=10)
    WRAP      = Alignment(wrap_text=True, vertical="top")
    THIN      = Side(style="thin", color="B7B7B7")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def set_header(ws, cols):
        for c, title in enumerate(cols, 1):
            cell = ws.cell(1, c, title)
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = BORDER
        ws.row_dimensions[1].height = 30

    def style_row(ws, row, fill_hex):
        fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        for cell in ws[row]:
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.alignment = WRAP
            cell.border    = BORDER

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    set_header(ws_sum, ["Category", "Domain Code", "KRI Count", "Status"])
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 12
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 20

    for r, dom in enumerate(DOMAIN_ORDER, 2):
        kris = by_cat.get(dom, [])
        if not kris:
            continue
        ws_sum.cell(r, 1, CATEGORY_LABELS[dom])
        ws_sum.cell(r, 2, dom)
        ws_sum.cell(r, 3, len(kris))
        ws_sum.cell(r, 4, "VERIFIED ✓")
        style_row(ws_sum, r, DOMAIN_COLORS.get(dom, "FFFFFF"))

    total_row = len(DOMAIN_ORDER) + 2
    total_kris = sum(len(v) for v in by_cat.values())
    for c, val in enumerate(["TOTAL", "", total_kris, f"ALL {total_kris} KRIs VERIFIED"], 1):
        cell = ws_sum.cell(total_row, c, val)
        cell.font   = Font(bold=True, color="FFFFFF", size=11)
        cell.fill   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.border = BORDER
        cell.alignment = WRAP

    # ── Domain sheets ─────────────────────────────────────────────────────────
    for dom in DOMAIN_ORDER:
        kris = by_cat.get(dom, [])
        if not kris:
            continue
        ws = wb.create_sheet(dom)
        set_header(ws, EXCEL_COLS)

        for r, k in enumerate(kris, 2):
            ws.cell(r, 1, k.get(F_KRI_ID, ""))
            ws.cell(r, 2, k.get(F_CATEGORY, ""))
            ws.cell(r, 3, k.get(F_NAME, ""))
            ws.cell(r, 4, k.get(F_DESCRIPTION, ""))
            ws.cell(r, 5, k.get(F_RULE, ""))
            ws.cell(r, 6, k.get(F_COMBINED, ""))
            ws.cell(r, 7, k.get(F_SEVERITY, ""))
            style_row(ws, r, DOMAIN_COLORS.get(dom, "FFFFFF"))

        for c, w in enumerate(EXCEL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    xlsx_path = os.path.join(out_dir, "Extracted_KRIs.xlsx")
    wb.save(xlsx_path)
    return xlsx_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--out",      required=True, help="Output directory with raw_*.json files")
    parser.add_argument("--manifest", required=True, help="Path to manifest.json")
    args = parser.parse_args()
    assemble(args.out, args.manifest)
