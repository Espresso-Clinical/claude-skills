"""
Step 4A — Assembly + Excel Generation
Merges raw_SOA.json, raw_ELIG.json, raw_SAF.json, raw_END.json, raw_OPS.json
into a single extracted_kris.json AND generates Extracted_KRIs.xlsx. No LLM required.

Usage:
  python step4a_assemble.py --out /path/to/output/ --manifest /path/to/manifest.json

Dependencies:
  pip install openpyxl --break-system-packages -q
"""
import json, os, re, argparse

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed. Excel generation skipped. Run: pip install openpyxl")

CATEGORY_LABELS = {
    "SOA": "Schedule of Activities",
    "ELIG": "Eligibility",
    "SAF": "Safety & Toxicity",
    "END": "Endpoints & Statistics",
    "OPS": "Operations & Compliance"
}

SHEET_CONFIG = [
    {"sheet_name": "SOA",         "cat_key": "SOA",  "has_footnotes": True},
    {"sheet_name": "ELIGIBILITY", "cat_key": "ELIG", "has_footnotes": False},
    {"sheet_name": "SAF&TOX",     "cat_key": "SAF",  "has_footnotes": False},
    {"sheet_name": "END&STAT",    "cat_key": "END",  "has_footnotes": False},
    {"sheet_name": "OPE&COM",     "cat_key": "OPS",  "has_footnotes": False},
]


def assemble(out_dir: str, manifest_path: str) -> str:
    """Merge raw category files into extracted_kris.json."""
    with open(manifest_path) as f:
        manifest = json.load(f)

    all_kris, categories_meta, by_cat = [], [], {}

    for cat in ["SOA", "ELIG", "SAF", "END", "OPS"]:
        raw_path = os.path.join(out_dir, f"raw_{cat}.json")
        if not os.path.exists(raw_path):
            print(f"  {cat}: not found, skipping")
            continue

        with open(raw_path) as f:
            data = json.load(f)

        # Handle multiple JSON formats
        if isinstance(data, list):
            kris = data
        else:
            kris = data.get("kris", data.get("KRIs", []))
            if not kris and isinstance(data, dict):
                for k, v in data.items():
                    if isinstance(v, list) and len(v) > 0:
                        kris = v
                        break

        # Normalize fields, deduplicate by rule_for_llm
        seen, cleaned = set(), []
        for k in kris:
            if not isinstance(k, dict):
                continue
            rule = (k.get("rule_for_llm") or "").strip().lower()
            if rule and rule in seen:
                continue
            if rule:
                seen.add(rule)
            if not k.get("rule_for_llm") and not k.get("kri_name"):
                continue
            entry = {
                "kri_id":              k.get("kri_id", ""),
                "kri_name":            k.get("kri_name", ""),
                "description":         k.get("description", None),
                "category_id":         cat,
                "category_label":      CATEGORY_LABELS[cat],
                "rule_for_llm":        k.get("rule_for_llm", None),
                "protocol_reference":  k.get("protocol_reference", None),
                "additional_footnotes": k.get("additional_footnotes", None)
            }
            cleaned.append(entry)

        all_kris.extend(cleaned)
        by_cat[cat] = cleaned
        categories_meta.append({
            "id": cat,
            "label": CATEGORY_LABELS[cat],
            "kri_count": len(cleaned),
            "kri_ids": [k["kri_id"] for k in cleaned]
        })
        print(f"  {cat}: {len(cleaned)} KRIs")

    output = {
        "_meta": {
            "version": "9.0",
            "protocol": manifest.get("protocol_id", "UNKNOWN"),
            "sponsor": manifest.get("sponsor", ""),
            "extracted_by": "protocol-kri-extractor",
            "total_kris": len(all_kris),
            "total_categories": len(categories_meta)
        },
        "categories": categories_meta,
        "kris": all_kris
    }

    out_path = os.path.join(out_dir, "extracted_kris.json")
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"\n  Total: {len(all_kris)} KRIs -> {out_path}")

    # Generate Excel
    if HAS_OPENPYXL:
        xlsx_path = generate_excel(out_dir, by_cat)
        print(f"  Excel: {xlsx_path}")
    else:
        print("  Excel: SKIPPED (openpyxl not installed)")

    return out_path


def generate_excel(out_dir: str, by_cat: dict) -> str:
    """Generate Extracted_KRIs.xlsx with 5 category sheets.

    Format: blue header row, white bold text, frozen row 1, text wrapping,
    thin borders, auto-width columns.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Styles
    hdr_font = Font(bold=True, size=11, color='FFFFFF')
    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    wrap = Alignment(wrap_text=True, vertical='top')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cfg in SHEET_CONFIG:
        ws = wb.create_sheet(cfg["sheet_name"])
        kris = by_cat.get(cfg["cat_key"], [])

        # Column headers
        if cfg["has_footnotes"]:
            headers = ["KRI Name", "Description", "Category",
                       "Rule to Check- for LLM", "Referance", "Additional Footnotes"]
            widths = [35, 50, 20, 65, 35, 50]
        else:
            headers = ["KRI Name", "Description", "Category",
                       "Referance", "Rule to Check- for LLM"]
            widths = [35, 50, 20, 35, 65]

        # Write header row
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = wrap
            c.border = border

        # Write KRI rows
        for i, kri in enumerate(kris, 2):
            name = kri.get("kri_name", kri.get("kri_id", ""))
            desc = kri.get("description", "")
            cat_label = kri.get("category_label", cfg["cat_key"])
            rule = kri.get("rule_for_llm", "")
            ref = kri.get("protocol_reference", "")
            fn = kri.get("additional_footnotes") or ""

            if cfg["has_footnotes"]:
                row_data = [name, desc, cat_label, rule, ref, fn]
            else:
                row_data = [name, desc, cat_label, ref, rule]

            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=i, column=col, value=str(val) if val else "")
                c.alignment = wrap
                c.border = border

        # Set column widths
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        # Freeze header row
        ws.freeze_panes = 'A2'

    xlsx_path = os.path.join(out_dir, "Extracted_KRIs.xlsx")
    wb.save(xlsx_path)
    return xlsx_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", required=True, help="Output directory with raw_*.json files")
    parser.add_argument("--manifest", required=True, help="Path to manifest.json")
    args = parser.parse_args()
    assemble(args.out, args.manifest)
